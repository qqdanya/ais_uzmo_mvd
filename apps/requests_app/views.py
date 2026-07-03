import csv
import json
import mimetypes
import os
import tempfile
import zipfile
from datetime import timedelta
from pathlib import Path, PurePosixPath
from types import SimpleNamespace

from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Min, Q, Sum
from django.http import FileResponse, Http404, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import capfirst
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.audit.models import AuditLog
from apps.audit.utils import serialize_instance, write_audit
from apps.directory.forms import TerritorialOrganPhotoFolderForm, TerritorialOrganPhotoForm
from apps.directory.models import Department, TerritorialOrgan, TerritorialOrganPhoto, TerritorialOrganPhotoFolder

from .forms import TmcRequestForm, form_for_table
from .models import (
    AntiTerrorMeasure,
    BuildingRepairRequest,
    CitsiziEquipment,
    FireDepartmentRequest,
    NeedStatus,
    RequestPhotoLink,
    RequestStatusHistory,
    TmcRequest,
    TmcRequestItem,
    VehicleFuelRequest,
    VehicleRepairRequest,
    TmcProduct,
    normalize_product_name,
    ACTIVE_NEED_STATUS_CHOICES,
)
from .permissions import can_manage_photo_asset, can_view, can_write, user_primary_department
from .registry import TABLES, TABLE_BY_KEY


def is_htmx(request):
    return request.headers.get("HX-Request") == "true"


def active_organs():
    return TerritorialOrgan.objects.filter(is_active=True, parent__isnull=True).prefetch_related("children")


def selected_organs_from_request(request, fallback_organ):
    raw_ids = request.GET.getlist("organ_ids")
    if not raw_ids and request.GET.get("organ_ids"):
        raw_ids = request.GET["organ_ids"].split(",")
    ids = [int(value) for value in raw_ids if str(value).isdigit()]
    if not ids:
        return [fallback_organ]
    organs = list(TerritorialOrgan.objects.filter(pk__in=ids, is_active=True, parent__isnull=True).order_by("order_number", "name"))
    allowed = [organ for organ in organs if can_view(request.user, organ)]
    return allowed or [fallback_organ]


def selected_organs_querystring(organs):
    query = QueryDict(mutable=True)
    for organ in organs:
        query.appendlist("organ_ids", str(organ.pk))
    return query.urlencode()


def photo_matches_query(photo, query):
    query_normalized = query.casefold()
    return query_normalized in photo.description.casefold() or query_normalized in photo.original_filename.casefold()


DEPARTMENT_ICONS = {
    "tmc": "bi-box-seam",
    "transport": "bi-truck",
    "fire": "bi-fire",
    "antiterror": "bi-shield-lock",
    "citsizi": "bi-router",
    "uoto": "bi-building",
}


@login_required
def dashboard(request):
    organs = active_organs()
    departments = list(Department.objects.filter(is_active=True))
    for department in departments:
        department.icon_class = DEPARTMENT_ICONS.get(department.slug, "bi-folder2-open")
    selected_organ = organs.first()
    selected_department = departments[0] if departments else None
    return render(request, "dashboard/index.html", {"organs": organs, "departments": departments, "selected_organ": selected_organ, "selected_department": selected_department, "tables": TABLES})


@login_required
def organ_info(request, pk):
    organ = get_object_or_404(TerritorialOrgan.objects.prefetch_related("children"), pk=pk, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    return render(request, "partials/organ_info.html", {"organ": organ})


@login_required
def department_tables(request, organ_id, department_slug):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    department = get_object_or_404(Department, slug=department_slug, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    selected_organs = selected_organs_from_request(request, organ)
    department_tables = TABLES[department.slug]
    requested_table_key = request.GET.get("table")
    table = next((item for item in department_tables if item["key"] == requested_table_key), department_tables[0])
    table_query = request.GET.copy()
    table_query.pop("table", None)
    return render(
        request,
        "partials/tables_panel.html",
        {
            "organ": selected_organs[0],
            "department": department,
            "tables": department_tables,
            "active_table": table,
            "selected_organs": selected_organs,
            "is_multi_organ": len(selected_organs) > 1,
            "organ_querystring": selected_organs_querystring(selected_organs) if len(selected_organs) > 1 else "",
            "table_querystring": table_query.urlencode(),
        },
    )


def filtered_queryset(request, table, organs):
    qs = table["model"].objects.select_related("territorial_organ", "created_by", "updated_by").filter(territorial_organ__in=organs, is_deleted=False)
    if table["key"] in REQUEST_TABLE_CONFIG:
        return request_table_queryset(request, table["key"], organs, include_status=True)
    if request.GET.get("equipment_type") and hasattr(table["model"], "equipment_type"):
        qs = qs.filter(equipment_type=request.GET["equipment_type"])
    if request.GET.get("status"):
        qs = qs.filter(status=request.GET["status"])
    return qs


STATE_SNAPSHOT_TABLES = {
    "fire-extinguishers",
    "fire-alarm",
    "security-alarm",
    "service-housing",
}

STATE_SNAPSHOT_MODE_CHOICES = (
    ("current", "Последняя запись"),
    ("history", "История записей"),
)


def state_snapshot_mode(request, table_key):
    if table_key not in STATE_SNAPSHOT_TABLES:
        return ""
    return "history" if request.GET.get("state_mode") == "history" else "current"


def state_snapshot_queryset(request, table_key, qs):
    if state_snapshot_mode(request, table_key) == "history":
        return qs
    latest_ids = []
    organ_ids = qs.values_list("territorial_organ_id", flat=True).distinct()
    for organ_id in organ_ids:
        latest_id = (
            qs.filter(territorial_organ_id=organ_id)
            .order_by("-state_date", "-created_at", "-pk")
            .values_list("pk", flat=True)
            .first()
        )
        if latest_id:
            latest_ids.append(latest_id)
    return qs.filter(pk__in=latest_ids).order_by("territorial_organ__name", "-state_date", "-created_at")


def request_date_filter_defaults(model, organs):
    oldest_date = model.objects.filter(territorial_organ__in=organs, is_deleted=False).aggregate(oldest=Min("request_date")).get("oldest")
    return {
        "date_from": oldest_date.isoformat() if oldest_date else "",
        "date_to": timezone.localdate().isoformat(),
    }


def request_date_filter_values(request, model, organs):
    defaults = request_date_filter_defaults(model, organs)
    date_from = request.GET.get("date_from") if "date_from" in request.GET else defaults["date_from"]
    date_to = request.GET.get("date_to") if "date_to" in request.GET else defaults["date_to"]
    return {"date_from": date_from, "date_to": date_to}


def request_table_date_filter_defaults(table_key, organs):
    return request_date_filter_defaults(REQUEST_TABLE_CONFIG[table_key]["model"], organs)


def request_table_date_filter_values(request, table_key, organs):
    return request_date_filter_values(request, REQUEST_TABLE_CONFIG[table_key]["model"], organs)


def related_search_values(obj, field_name):
    values = [obj]
    for part in field_name.split("__"):
        next_values = []
        for value in values:
            attr = getattr(value, part, None)
            if attr is None:
                continue
            if hasattr(attr, "all") and callable(attr.all):
                next_values.extend(attr.all())
            else:
                next_values.append(attr)
        values = next_values
    return values


def object_matches_casefold_search(obj, search_fields, query):
    query = query.casefold()
    for field_name in search_fields:
        for value in related_search_values(obj, field_name):
            if query in str(value or "").casefold():
                return True
    return False


def apply_casefold_search(qs, search_fields, query):
    query = query.strip()
    if not query:
        return qs
    matched_ids = [obj.pk for obj in qs if object_matches_casefold_search(obj, search_fields, query)]
    if not matched_ids:
        return qs.none()
    return qs.filter(pk__in=matched_ids)


def request_table_queryset(request, table_key, organs, include_status=False):
    config = REQUEST_TABLE_CONFIG[table_key]
    qs = config["model"].objects.select_related("territorial_organ", "created_by", "updated_by")
    if config.get("prefetch"):
        qs = qs.prefetch_related(*config["prefetch"])
    qs = qs.filter(territorial_organ__in=organs, is_deleted=False)

    date_filters = request_table_date_filter_values(request, table_key, organs)
    date_from = parse_date(date_filters["date_from"])
    date_to = parse_date(date_filters["date_to"])
    if date_from:
        qs = qs.filter(request_date__gte=date_from)
    if date_to:
        qs = qs.filter(request_date__lte=date_to)
    if config.get("equipment_type_filter") and valid_equipment_type(request.GET.get("equipment_type")):
        qs = qs.filter(equipment_type=request.GET["equipment_type"])
    qs = apply_casefold_search(qs, config["search_fields"], request.GET.get("q", ""))
    if include_status and request.GET.get("status") in dict(ACTIVE_NEED_STATUS_CHOICES):
        qs = qs.filter(status=request.GET["status"])
    return qs


def request_status_stats(qs):
    return {
        "in_work_count": qs.filter(status=NeedStatus.IN_WORK).count(),
        "done_count": qs.filter(status=NeedStatus.DONE).count(),
        "rejected_count": qs.filter(status=NeedStatus.REJECTED).count(),
    }


def format_filter_date(value):
    date = parse_date(value or "")
    return date.strftime("%d.%m.%Y") if date else value


def active_table_conditions(request, table_key, selected_organs, group_mode="requests"):
    conditions = []
    if len(selected_organs) > 1:
        conditions.append(f"выборочно: {len(selected_organs)} органов")
    if group_mode == "products":
        conditions.append("группировка: По ТМЦ")
    if group_mode == "organs":
        conditions.append("группировка: По территориальному органу")
    if group_mode == "dates":
        conditions.append("группировка: По дате")
    query = request.GET.get("q", "").strip()
    if query:
        conditions.append(f"поиск: {query}")
    status_labels = dict(ACTIVE_NEED_STATUS_CHOICES)
    status = request.GET.get("status")
    if status in status_labels:
        conditions.append(f"исполнение: {status_labels[status]}")
    if table_key == "citsizi-equipment":
        equipment_labels = dict(CitsiziEquipment._meta.get_field("equipment_type").choices)
        equipment_type = request.GET.get("equipment_type")
        if equipment_type in equipment_labels:
            conditions.append(f"тип техники: {equipment_labels[equipment_type]}")
    if request.GET.get("date_from"):
        conditions.append(f"с {format_filter_date(request.GET['date_from'])}")
    if request.GET.get("date_to"):
        conditions.append(f"по {format_filter_date(request.GET['date_to'])}")
    return conditions


FIRE_EXTINGUISHER_SOON_DAYS = 30
FIRE_EXTINGUISHER_EXPIRY_STATE_CHOICES = (
    ("", "Все сроки"),
    ("valid", "Годные"),
    ("soon", "Скоро истекает"),
    ("expired", "Истекшие"),
)
FIRE_EXTINGUISHER_EXPIRY_ORDER_CHOICES = (
    ("", "По порядку добавления"),
    ("soonest", "Сначала истекающие"),
    ("latest", "Сначала с большим сроком"),
)


def fire_extinguisher_expiry_window():
    today = timezone.localdate()
    return today, today + timedelta(days=FIRE_EXTINGUISHER_SOON_DAYS)


def fire_extinguisher_filtered_queryset(request, qs):
    today, soon_until = fire_extinguisher_expiry_window()
    expiry_state = request.GET.get("expiry_state", "")
    if expiry_state == "expired":
        qs = qs.filter(expiry_date__lt=today)
    elif expiry_state == "soon":
        qs = qs.filter(expiry_date__gte=today, expiry_date__lte=soon_until)
    elif expiry_state == "valid":
        qs = qs.filter(expiry_date__gt=soon_until)

    expiry_from = parse_date(request.GET.get("expiry_from", ""))
    expiry_to = parse_date(request.GET.get("expiry_to", ""))
    if expiry_from:
        qs = qs.filter(expiry_date__gte=expiry_from)
    if expiry_to:
        qs = qs.filter(expiry_date__lte=expiry_to)

    expiry_order = request.GET.get("expiry_order", "")
    if expiry_order == "latest":
        return qs.order_by("-expiry_date", "-state_date", "-created_at")
    if expiry_order == "soonest":
        return qs.order_by("expiry_date", "-state_date", "-created_at")
    return qs.order_by("-created_at", "-id")


def fire_extinguisher_active_conditions(request, selected_organs):
    conditions = []
    if len(selected_organs) > 1:
        conditions.append(f"выборочно: {len(selected_organs)} органов")
    expiry_state_labels = dict(FIRE_EXTINGUISHER_EXPIRY_STATE_CHOICES)
    expiry_state = request.GET.get("expiry_state", "")
    if expiry_state:
        conditions.append(f"срок: {expiry_state_labels.get(expiry_state, expiry_state)}")
    expiry_order_labels = dict(FIRE_EXTINGUISHER_EXPIRY_ORDER_CHOICES)
    expiry_order = request.GET.get("expiry_order", "")
    if expiry_order:
        conditions.append(f"сортировка: {expiry_order_labels.get(expiry_order, expiry_order)}")
    if request.GET.get("expiry_from"):
        conditions.append(f"срок с {format_filter_date(request.GET['expiry_from'])}")
    if request.GET.get("expiry_to"):
        conditions.append(f"срок по {format_filter_date(request.GET['expiry_to'])}")
    return conditions


def tmc_grouped_rows(qs):
    return (
        TmcRequestItem.objects.filter(request__in=qs)
        .values("product_id", "product__name", "name", "unit")
        .annotate(
            request_count=Count("request_id", distinct=True),
            organ_count=Count("request__territorial_organ_id", distinct=True),
            total_quantity=Sum("quantity"),
        )
        .order_by("-request_count", "-total_quantity", "product__name", "name", "unit")
    )


def tmc_organ_grouped_rows(qs):
    rows = request_organ_grouped_rows(qs)
    quantities = {
        row["request__territorial_organ_id"]: row
        for row in TmcRequestItem.objects.filter(request__in=qs)
        .values("request__territorial_organ_id", "request__territorial_organ__name")
        .annotate(
            request_count=Count("request_id", distinct=True),
            position_count=Count("id"),
            total_quantity=Sum("quantity"),
        )
    }
    for row in rows:
        item_row = quantities.get(row["territorial_organ_id"], {})
        row["request__territorial_organ__name"] = row.get("territorial_organ__name")
        row["position_count"] = item_row.get("position_count") or 0
        row["total_quantity"] = item_row.get("total_quantity") or 0
    return rows


def tmc_date_grouped_rows(qs):
    rows = request_date_grouped_rows(qs)
    quantities = {
        row["request__request_date"]: row
        for row in TmcRequestItem.objects.filter(request__in=qs)
        .values("request__request_date")
        .annotate(position_count=Count("id"), total_quantity=Sum("quantity"))
    }
    for row in rows:
        item_row = quantities.get(row["request_date"], {})
        row["position_count"] = item_row.get("position_count") or 0
        row["total_quantity"] = item_row.get("total_quantity") or 0
    return rows


def request_date_grouped_rows(qs):
    return list(
        qs.values("request_date")
        .annotate(
            request_count=Count("id"),
            organ_count=Count("territorial_organ_id", distinct=True),
            in_work_count=Count("id", filter=Q(status=NeedStatus.IN_WORK)),
            done_count=Count("id", filter=Q(status=NeedStatus.DONE)),
            rejected_count=Count("id", filter=Q(status=NeedStatus.REJECTED)),
        )
        .order_by("-request_date")
    )


def request_organ_grouped_rows(qs):
    return list(
        qs.values("territorial_organ_id", "territorial_organ__name")
        .annotate(
            request_count=Count("id"),
            in_work_count=Count("id", filter=Q(status=NeedStatus.IN_WORK)),
            done_count=Count("id", filter=Q(status=NeedStatus.DONE)),
            rejected_count=Count("id", filter=Q(status=NeedStatus.REJECTED)),
        )
        .order_by("territorial_organ__name")
    )


def tmc_grouped_summary(qs, grouped_count):
    items = TmcRequestItem.objects.filter(request__in=qs)
    return {
        "position_count": grouped_count,
        "request_count": qs.count(),
        "organ_count": qs.values("territorial_organ_id").distinct().count(),
        "total_quantity": items.aggregate(total=Sum("quantity")).get("total") or 0,
    }


def tmc_organ_grouped_summary(qs, grouped_count):
    items = TmcRequestItem.objects.filter(request__in=qs)
    return {
        "organ_count": grouped_count,
        "request_count": qs.count(),
        "position_count": items.count(),
        "total_quantity": items.aggregate(total=Sum("quantity")).get("total") or 0,
    }


def tmc_date_grouped_summary(qs, grouped_count):
    items = TmcRequestItem.objects.filter(request__in=qs)
    summary = request_grouped_summary(qs, date_count=grouped_count)
    summary.update({
        "date_count": grouped_count,
        "position_count": items.count(),
        "total_quantity": items.aggregate(total=Sum("quantity")).get("total") or 0,
    })
    return summary


def request_grouped_summary(qs, date_count=None, organ_count=None):
    summary = request_status_stats(qs)
    summary.update(
        {
            "request_count": qs.count(),
            "organ_count": organ_count if organ_count is not None else qs.values("territorial_organ_id").distinct().count(),
        }
    )
    if date_count is not None:
        summary["date_count"] = date_count
    return summary


def attach_tmc_drilldown_querystrings(rows, base_querystring):
    for row in rows:
        product_name = row.get("product__name") or row.get("name") or ""
        drilldown_querystring = base_querystring.copy()
        drilldown_querystring["q"] = product_name
        row["drilldown_querystring"] = drilldown_querystring.urlencode()
    return rows


def request_group_mode(request, table_key, is_multi_organ):
    group = request.GET.get("group")
    if table_key == "tmc-requests" and group == "products":
        return "products"
    if group == "organs" and is_multi_organ:
        return "organs"
    if group == "dates":
        return "dates"
    return "requests"


def table_view_query_fields(querystring):
    fields = []
    for name, values in querystring.lists():
        if name in {"page", "group", "state_mode"}:
            continue
        fields.extend({"name": name, "value": value} for value in values)
    return fields


def row_count(rows):
    return len(rows) if isinstance(rows, list) else rows.count()


def valid_equipment_type(value):
    return value in {choice[0] for choice in CitsiziEquipment._meta.get_field("equipment_type").choices}


STATUS_HISTORY_TABLES = {
    "tmc-requests",
    "anti-terror",
    "building-repair",
    "citsizi-equipment",
    "vehicle-repair",
    "vehicle-fuel",
    "fire-requests",
}


COMPLETED_DATE_FIELDS = {
    "citsizi-equipment": "due_date",
    "tmc-requests": "due_date",
}


REQUEST_TABLE_CONFIG = {
    "tmc-requests": {
        "model": TmcRequest,
        "search_fields": ("request_number", "comment", "items__name"),
        "prefetch": ("items",),
        "distinct_search": True,
        "completed_label": "Дата исполнения",
    },
    "vehicle-repair": {
        "model": VehicleRepairRequest,
        "search_fields": ("request_number", "comment"),
        "completed_label": "Дата исполнения заявки",
    },
    "vehicle-fuel": {
        "model": VehicleFuelRequest,
        "search_fields": ("request_number", "comment"),
        "completed_label": "Дата исполнения заявки",
    },
    "fire-requests": {
        "model": FireDepartmentRequest,
        "search_fields": ("request_number", "comment"),
        "completed_label": "Дата исполнения заявки",
    },
    "anti-terror": {
        "model": AntiTerrorMeasure,
        "search_fields": ("request_number", "comment"),
        "completed_label": "Дата исполнения заявки",
    },
    "citsizi-equipment": {
        "model": CitsiziEquipment,
        "search_fields": ("request_number", "comment"),
        "equipment_type_filter": True,
        "completed_label": "Дата исполнения заявки",
    },
    "building-repair": {
        "model": BuildingRepairRequest,
        "search_fields": ("request_number", "comment"),
        "completed_label": "Дата исполнения заявки",
    },
}

REQUEST_PHOTO_TABLES = set(REQUEST_TABLE_CONFIG)
REQUEST_PHOTO_PICKER_PAGE_SIZE = 12
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOWNLOAD_READY_COOKIE_PREFIX = "download-ready-"
SIMPLE_REQUEST_XLSX_CONFIG = {
    "widths": {
        "request_number": 18,
        "request_date": 14,
        "status": 22,
        "comment": 38,
    },
    "center_columns": {"request_number", "request_date", "status"},
}


class TemporaryDownloadFile:
    def __init__(self, path):
        self.path = path
        self.file = open(path, "rb")

    def __getattr__(self, name):
        return getattr(self.file, name)

    def close(self):
        try:
            self.file.close()
        finally:
            try:
                os.remove(self.path)
            except FileNotFoundError:
                pass


def temporary_download_response(path, filename, content_type):
    return FileResponse(TemporaryDownloadFile(path), as_attachment=True, filename=filename, content_type=content_type)


def download_ready_response(request, response):
    token = request.GET.get("download_token", "").strip()
    if token and all(char.isalnum() or char in "-_" for char in token):
        response.set_cookie(f"{DOWNLOAD_READY_COOKIE_PREFIX}{token}", "1", max_age=120, path="/", samesite="Lax")
    return response


def workbook_file_response(workbook, filename):
    temp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = temp.name
    temp.close()
    try:
        workbook.save(path)
    except Exception:
        try:
            os.remove(path)
        except FileNotFoundError:
            pass
        raise
    return temporary_download_response(path, filename, XLSX_CONTENT_TYPE)


def csv_file_response(filename, rows):
    temp = tempfile.NamedTemporaryFile("w", suffix=".csv", newline="", encoding="utf-8-sig", delete=False)
    path = temp.name
    try:
        writer = csv.writer(temp)
        writer.writerows(rows)
    except Exception:
        temp.close()
        try:
            os.remove(path)
        except FileNotFoundError:
            pass
        raise
    temp.close()
    return temporary_download_response(path, filename, "text/csv; charset=utf-8")


def request_content_type_for_model(model):
    return ContentType.objects.get_for_model(model, for_concrete_model=False)


def request_content_type_for_object(obj):
    return ContentType.objects.get_for_model(obj, for_concrete_model=False)


def available_request_photos(organ):
    return (
        organ.photos.select_related("folder", "created_by")
        .filter(is_deleted=False)
        .filter(Q(folder__isnull=True) | Q(folder__is_deleted=False))
        .order_by("-created_at", "-pk")
    )


def selected_request_photo_ids(instance):
    if not instance:
        return []
    content_type = request_content_type_for_object(instance)
    return list(
        RequestPhotoLink.objects.filter(
            territorial_organ=instance.territorial_organ,
            content_type=content_type,
            object_id=instance.pk,
            photo__is_deleted=False,
        )
        .filter(Q(photo__folder__isnull=True) | Q(photo__folder__is_deleted=False))
        .values_list("photo_id", flat=True)
    )


def request_photo_picker_context(request, organ, selected_ids):
    selected_ids = {int(value) for value in selected_ids if str(value).isdigit()}
    query = request.GET.get("photo_q", "").strip()
    folder = request.GET.get("photo_folder", "").strip()
    sort = request.GET.get("photo_sort", "newest")
    folders = organ.photo_folders.filter(is_deleted=False).order_by("name")

    selected_photos = list(available_request_photos(organ).filter(pk__in=selected_ids))
    qs = available_request_photos(organ).exclude(pk__in=selected_ids)
    if folder == "__root__":
        qs = qs.filter(folder__isnull=True)
    elif folder.isdigit():
        qs = qs.filter(folder_id=folder)
    if query:
        query_normalized = query.casefold()
        qs = [photo for photo in qs if query_normalized in photo.description.casefold() or query_normalized in photo.original_filename.casefold()]
        qs = sorted(qs, key=lambda photo: (photo.created_at, photo.pk), reverse=sort != "oldest")
    else:
        qs = qs.order_by("created_at", "pk") if sort == "oldest" else qs.order_by("-created_at", "-pk")

    page = Paginator(qs, REQUEST_PHOTO_PICKER_PAGE_SIZE).get_page(request.GET.get("photo_page"))
    photos = selected_photos + list(page.object_list)
    for photo in photos:
        photo.is_attached_to_request = photo.pk in selected_ids
    return {
        "available_photos": photos,
        "attached_photo_ids": selected_ids,
        "attached_photo_count": len(selected_ids),
        "photo_picker_page": page,
        "photo_picker_page_links": page.paginator.get_elided_page_range(page.number, on_each_side=1, on_ends=1),
        "photo_picker_folders": folders,
        "photo_picker_query": query,
        "photo_picker_folder": folder,
        "photo_picker_sort": sort,
    }


def request_photo_form_context(request, organ, selected_ids):
    return request_photo_picker_context(request, organ, selected_ids)


def sync_request_photos(obj, photo_ids, user):
    selected_ids = {int(value) for value in photo_ids if str(value).isdigit()}
    content_type = request_content_type_for_object(obj)
    valid_photo_ids = set(available_request_photos(obj.territorial_organ).filter(pk__in=selected_ids).values_list("pk", flat=True))
    links = RequestPhotoLink.objects.filter(territorial_organ=obj.territorial_organ, content_type=content_type, object_id=obj.pk)
    existing_ids = set(links.values_list("photo_id", flat=True))
    removed_ids = existing_ids - valid_photo_ids
    links.exclude(photo_id__in=valid_photo_ids).delete()
    added_ids = valid_photo_ids - existing_ids
    RequestPhotoLink.objects.bulk_create(
        [
            RequestPhotoLink(
                territorial_organ=obj.territorial_organ,
                photo_id=photo_id,
                content_type=content_type,
                object_id=obj.pk,
                created_by=user,
            )
            for photo_id in added_ids
        ],
        ignore_conflicts=True,
    )
    return {"added": added_ids, "removed": removed_ids}


def photo_names_for_audit(photo_ids):
    if not photo_ids:
        return ""
    names = list(
        TerritorialOrganPhoto.objects.filter(pk__in=photo_ids)
        .order_by("original_filename", "pk")
        .values_list("original_filename", flat=True)
    )
    return ", ".join(name or "фотография" for name in names)


def write_request_photo_audit_events(obj, changes, request):
    added = changes.get("added") or set()
    removed = changes.get("removed") or set()
    if added:
        write_audit(
            AuditLog.Action.UPDATE,
            obj,
            old_values={"photos": ""},
            new_values={"audit_event": "request_photos_attached", "photos": photo_names_for_audit(added)},
            request=request,
        )
    if removed:
        write_audit(
            AuditLog.Action.UPDATE,
            obj,
            old_values={"photos": photo_names_for_audit(removed)},
            new_values={"audit_event": "request_photos_detached", "photos": ""},
            request=request,
        )


def write_status_change_audit_event(obj, old_status, new_status, request):
    write_audit(
        AuditLog.Action.UPDATE,
        obj,
        old_values={"status": old_status},
        new_values={"audit_event": "request_status_changed", "status": new_status},
        request=request,
    )


def attach_request_photo_counts(objects, model, organs):
    objects = list(objects)
    if not objects:
        return
    content_type = request_content_type_for_model(model)
    counts = dict(
        RequestPhotoLink.objects.filter(
            territorial_organ__in=organs,
            content_type=content_type,
            object_id__in=[obj.pk for obj in objects],
            photo__is_deleted=False,
        )
        .filter(Q(photo__folder__isnull=True) | Q(photo__folder__is_deleted=False))
        .values("object_id")
        .annotate(total=Count("id"))
        .values_list("object_id", "total")
    )
    for obj in objects:
        obj.attached_photo_count = counts.get(obj.pk, 0)


XLSX_EXPORT_CONFIG = {
    "vehicle-inventory": {
        "widths": {
            "state_date": 14,
            "required_count": 14,
            "available_count": 14,
            "broken_count": 16,
            "writeoff_count": 38,
        },
        "center_columns": {"state_date", "required_count", "available_count", "broken_count", "writeoff_count"},
    },
    "vehicle-repair": {
        **SIMPLE_REQUEST_XLSX_CONFIG,
    },
    "vehicle-fuel": {
        **SIMPLE_REQUEST_XLSX_CONFIG,
    },
    "fire-extinguishers": {
        "widths": {
            "state_date": 14,
            "required_count": 14,
            "available_count": 14,
            "expiry_date": 24,
            "writeoff_count": 18,
        },
        "center_columns": {"state_date", "required_count", "available_count", "expiry_date", "writeoff_count"},
    },
    "fire-alarm": {
        "widths": {
            "state_date": 14,
            "required_objects": 28,
            "equipped_objects": 26,
            "broken_objects": 28,
        },
        "center_columns": {"state_date", "required_objects", "equipped_objects", "broken_objects"},
    },
    "security-alarm": {
        "widths": {
            "state_date": 14,
            "required_objects": 28,
            "equipped_objects": 26,
            "broken_objects": 28,
        },
        "center_columns": {"state_date", "required_objects", "equipped_objects", "broken_objects"},
    },
    "fire-requests": {
        "widths": {
            "request_number": 18,
            "request_date": 14,
            "status": 22,
            "comment": 38,
        },
        "center_columns": {"request_number", "request_date", "status"},
    },
    "anti-terror": {
        "widths": {
            "request_number": 18,
            "request_date": 14,
            "status": 22,
            "comment": 38,
        },
        "center_columns": {"request_number", "request_date", "status"},
    },
    "citsizi-equipment": {
        "widths": {
            "request_number": 18,
            "request_date": 14,
            "quantity": 14,
            "status": 22,
            "equipment_type": 28,
            "comment": 38,
        },
        "center_columns": {"request_number", "request_date", "quantity", "status", "equipment_type"},
    },
    "service-housing": {
        "widths": {
            "state_date": 14,
            "total_count": 18,
            "used_by_staff": 24,
            "ready_to_move": 20,
        },
        "center_columns": {"state_date", "total_count", "used_by_staff", "ready_to_move"},
    },
    "building-repair": {
        "widths": {
            "request_number": 18,
            "request_date": 14,
            "status": 22,
            "comment": 38,
        },
        "center_columns": {"request_number", "request_date", "status"},
    },
}


def completed_date_field(table_key):
    return COMPLETED_DATE_FIELDS.get(table_key, "completed_at")


def status_history_content_type(obj):
    return ContentType.objects.get_for_model(obj, for_concrete_model=False)


def status_history_queryset(obj):
    return RequestStatusHistory.objects.select_related("changed_by").filter(content_type=status_history_content_type(obj), object_id=obj.pk)


def attach_status_history_flags(objects, model):
    object_ids = [obj.pk for obj in objects]
    if not object_ids:
        return
    content_type = ContentType.objects.get_for_model(model, for_concrete_model=False)
    history_ids = set(
        RequestStatusHistory.objects.filter(content_type=content_type, object_id__in=object_ids).values_list("object_id", flat=True)
    )
    for obj in objects:
        obj.has_status_history_entries = obj.pk in history_ids


def create_status_history(obj, old_status, new_status, completed_at, changed_by, note):
    return RequestStatusHistory.objects.create(
        content_type=status_history_content_type(obj),
        object_id=obj.pk,
        old_status=old_status,
        new_status=new_status,
        completed_at=completed_at,
        changed_by=changed_by,
        note=note,
    )


def display_fields(table):
    field_names = table["fields"]
    fields_by_name = {field.name: field for field in table["model"]._meta.fields}
    computed_labels = {"items_summary": "наименования"}
    return [fields_by_name.get(name) or SimpleNamespace(name=name, verbose_name=computed_labels.get(name, name)) for name in field_names]


def table_header_labels(fields):
    return [capfirst(field.verbose_name) for field in fields]


def export_cell_value(obj, field):
    display = getattr(obj, f"get_{field.name}_display", None)
    value = display() if callable(display) else getattr(obj, field.name)
    if hasattr(value, "strftime") and getattr(field, "get_internal_type", lambda: "")() == "DateField":
        return value.strftime("%d.%m.%Y")
    return value


@login_required
def table_data(request, organ_id, table_key):
    table = TABLE_BY_KEY[table_key]
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    selected_organs = selected_organs_from_request(request, organ)
    is_multi_organ = len(selected_organs) > 1
    table_stats = {}
    table_filters = {}
    table_filter_defaults = {}
    qs = filtered_queryset(request, table, selected_organs)
    is_request_table = table_key in REQUEST_TABLE_CONFIG
    is_fire_extinguisher_table = table_key == "fire-extinguishers"
    is_state_snapshot_table = table_key in STATE_SNAPSHOT_TABLES
    current_state_mode = state_snapshot_mode(request, table_key)
    current_group_mode = request_group_mode(request, table_key, is_multi_organ) if is_request_table else "requests"
    if is_state_snapshot_table:
        qs = state_snapshot_queryset(request, table_key, qs)
    if is_fire_extinguisher_table:
        qs = fire_extinguisher_filtered_queryset(request, qs)
    is_request_grouped = current_group_mode in {"products", "organs", "dates"}
    is_tmc_grouped = table_key == "tmc-requests" and is_request_grouped
    is_tmc_product_grouped = table_key == "tmc-requests" and current_group_mode == "products"
    is_organ_grouped = current_group_mode == "organs"
    is_date_grouped = current_group_mode == "dates"
    if is_request_table:
        table_filter_defaults = request_table_date_filter_defaults(table_key, selected_organs)
        table_filters = request_table_date_filter_values(request, table_key, selected_organs)
        stats_qs = request_table_queryset(request, table_key, selected_organs)
        table_stats = request_status_stats(stats_qs)
    if is_tmc_product_grouped:
        page_qs = tmc_grouped_rows(qs)
    elif is_organ_grouped:
        page_qs = tmc_organ_grouped_rows(qs) if table_key == "tmc-requests" else request_organ_grouped_rows(qs)
    elif is_date_grouped:
        page_qs = tmc_date_grouped_rows(qs) if table_key == "tmc-requests" else request_date_grouped_rows(qs)
    else:
        page_qs = qs
    grouped_summary = {}
    grouped_count = row_count(page_qs)
    if is_tmc_product_grouped:
        grouped_summary = tmc_grouped_summary(qs, grouped_count)
    elif is_organ_grouped:
        grouped_summary = tmc_organ_grouped_summary(qs, grouped_count) if table_key == "tmc-requests" else request_grouped_summary(qs, organ_count=grouped_count)
    elif is_date_grouped:
        grouped_summary = tmc_date_grouped_summary(qs, grouped_count) if table_key == "tmc-requests" else request_grouped_summary(qs, date_count=grouped_count)
    paginator = Paginator(page_qs, 20)
    page = paginator.get_page(request.GET.get("page"))
    if table_key in REQUEST_PHOTO_TABLES and not is_request_grouped:
        attach_request_photo_counts(page.object_list, table["model"], selected_organs)
    if table_key in STATUS_HISTORY_TABLES and not is_request_grouped:
        attach_status_history_flags(page.object_list, table["model"])
    querystring = request.GET.copy()
    querystring.pop("page", None)
    list_querystring = querystring.copy()
    list_querystring.pop("group", None)
    grouped_querystring = querystring.copy()
    grouped_querystring["group"] = "products"
    organ_grouped_querystring = querystring.copy()
    organ_grouped_querystring["group"] = "organs"
    if is_tmc_product_grouped:
        page.object_list = attach_tmc_drilldown_querystrings(list(page.object_list), list_querystring)
    active_conditions = (
        fire_extinguisher_active_conditions(request, selected_organs)
        if is_fire_extinguisher_table
        else active_table_conditions(request, table_key, selected_organs, current_group_mode)
    )
    if is_state_snapshot_table and current_state_mode == "history":
        active_conditions.append("режим: История записей")
    writable_organ_ids = [selected_organ.pk for selected_organ in selected_organs if can_write(request.user, selected_organ, table["department"])]
    return render(
        request,
        "partials/table_data.html",
        {
            "organ": organ,
            "table": table,
            "fields": display_fields(table),
            "page": page,
            "table_page_links": page.paginator.get_elided_page_range(page.number, on_each_side=1, on_ends=1),
            "can_add": can_write(request.user, organ, table["department"]) and not is_multi_organ,
            "writable_organ_ids": writable_organ_ids,
            "table_querystring": querystring.urlencode(),
            "list_querystring": list_querystring.urlencode(),
            "grouped_querystring": grouped_querystring.urlencode(),
            "organ_grouped_querystring": organ_grouped_querystring.urlencode(),
            "table_view_query_fields": table_view_query_fields(querystring),
            "organ_querystring": selected_organs_querystring(selected_organs) if is_multi_organ else "",
            "status_choices": ACTIVE_NEED_STATUS_CHOICES,
            "table_stats": table_stats,
            "table_filters": table_filters,
            "table_filter_defaults": table_filter_defaults,
            "active_conditions": active_conditions,
            "grouped_summary": grouped_summary,
            "tmc_summary": grouped_summary,
            "is_request_table": is_request_table,
            "is_fire_extinguisher_table": is_fire_extinguisher_table,
            "is_state_snapshot_table": is_state_snapshot_table,
            "state_snapshot_mode": current_state_mode,
            "state_snapshot_mode_choices": STATE_SNAPSHOT_MODE_CHOICES,
            "is_request_grouped": is_request_grouped,
            "is_tmc_grouped": is_tmc_grouped,
            "is_tmc_product_grouped": is_tmc_product_grouped,
            "is_tmc_organ_grouped": table_key == "tmc-requests" and is_organ_grouped,
            "is_tmc_date_grouped": table_key == "tmc-requests" and is_date_grouped,
            "is_organ_grouped": is_organ_grouped,
            "is_date_grouped": is_date_grouped,
            "tmc_group_mode": current_group_mode,
            "group_mode": current_group_mode,
            "record_label": "позиций" if is_tmc_product_grouped else "органов" if is_organ_grouped else "дней" if is_date_grouped else "записей",
            "has_status_history": table_key in STATUS_HISTORY_TABLES,
            "search_placeholder": "Поиск по заявке и ТМЦ" if table_key == "tmc-requests" else "Поиск по заявке и описанию",
            "equipment_type_choices": CitsiziEquipment._meta.get_field("equipment_type").choices,
            "fire_extinguisher_expiry_state_choices": FIRE_EXTINGUISHER_EXPIRY_STATE_CHOICES,
            "fire_extinguisher_expiry_order_choices": FIRE_EXTINGUISHER_EXPIRY_ORDER_CHOICES,
            "selected_organs": selected_organs,
            "is_multi_organ": is_multi_organ,
        },
    )


def htmx_triggers(message, level="success"):
    return json.dumps({"modal:close": True, "toast": {"message": message, "level": level}})


def clean_product_name(value):
    return " ".join((value or "").split())


def product_tokens(value):
    return {token for token in normalize_product_name(value).split() if token}


def levenshtein_distance(left, right):
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    previous = list(range(len(right) + 1))
    for left_index, left_char in enumerate(left, start=1):
        current = [left_index]
        for right_index, right_char in enumerate(right, start=1):
            current.append(
                min(
                    previous[right_index] + 1,
                    current[right_index - 1] + 1,
                    previous[right_index - 1] + (left_char != right_char),
                )
            )
        previous = current
    return previous[-1]


def similarity_ratio(left, right):
    left = normalize_product_name(left)
    right = normalize_product_name(right)
    longest = max(len(left), len(right))
    if not longest:
        return 1
    return 1 - (levenshtein_distance(left, right) / longest)


def fuzzy_threshold(value):
    length = len(normalize_product_name(value))
    if length <= 4:
        return .92
    if length <= 7:
        return .82
    return .74


def best_fuzzy_similarity(query_normalized, product):
    candidates = [product.normalized_name]
    product_tokens_sorted = sorted(product_tokens(product.name))
    if len(product_tokens_sorted) > 1:
        candidates.append(" ".join(product_tokens_sorted))
    return max(similarity_ratio(query_normalized, candidate) for candidate in candidates if candidate)


def tmc_product_suggestions(query, limit=8):
    query = clean_product_name(query)
    if not query:
        return []
    query_normalized = normalize_product_name(query)
    query_tokens = product_tokens(query)
    suggestions = []
    for product in TmcProduct.objects.filter(is_active=True):
        product_tokens_set = product_tokens(product.name)
        if not product_tokens_set:
            continue
        if product.normalized_name == query_normalized:
            score = 100
        elif query_tokens and query_tokens.issubset(product_tokens_set):
            score = 90
        elif query_tokens and product_tokens_set.issubset(query_tokens):
            score = 80
        elif query_normalized and query_normalized in product.normalized_name:
            score = 70
        else:
            common_tokens = query_tokens & product_tokens_set
            if common_tokens:
                score = 50 + len(common_tokens)
            else:
                ratio = best_fuzzy_similarity(query_normalized, product)
                score = 40 + int(ratio * 10) if ratio >= fuzzy_threshold(query_normalized) else 0
        if score:
            suggestions.append((score, product.name.casefold(), product))
    suggestions.sort(key=lambda item: (-item[0], item[1]))
    return [product for _, __, product in suggestions[:limit]]


def get_or_create_tmc_product(name, unit, product_id=None):
    name = clean_product_name(name)
    unit = clean_product_name(unit) or "шт."
    if product_id and str(product_id).isdigit():
        product = TmcProduct.objects.filter(pk=product_id, is_active=True).first()
        if product:
            return product, False
    normalized_name = normalize_product_name(name)
    product = TmcProduct.objects.filter(normalized_name=normalized_name).first()
    if product:
        return product, False
    return TmcProduct.objects.create(name=name, unit=unit), True


def tmc_item_rows_from_request(request):
    rows = []
    errors = []
    product_ids = request.POST.getlist("item_product")
    names = request.POST.getlist("item_name")
    quantities = request.POST.getlist("item_quantity")
    units = request.POST.getlist("item_unit")
    for index, name in enumerate(names):
        name = clean_product_name(name)
        quantity_raw = quantities[index].strip() if index < len(quantities) else ""
        unit = clean_product_name(units[index]) if index < len(units) else "шт."
        product_id = product_ids[index].strip() if index < len(product_ids) else ""
        if not name and not quantity_raw:
            continue
        row = {"product_id": product_id, "name": name, "quantity": quantity_raw, "unit": unit or "шт."}
        if not name:
            errors.append("Укажите наименование в каждой заполненной позиции.")
        try:
            quantity = int(quantity_raw)
            if quantity <= 0:
                raise ValueError
            row["quantity"] = quantity
        except (TypeError, ValueError):
            errors.append(f"Укажите положительное количество для позиции «{name or 'без наименования'}».")
        rows.append(row)
    if not rows:
        errors.append("Добавьте хотя бы одну позицию заявки.")
        rows.append(tmc_blank_item_row())
    return rows, errors


def tmc_blank_item_row():
    return {"product_id": "", "name": "", "quantity": "", "unit": "шт."}


def tmc_item_rows_from_instance(instance):
    if not instance:
        return [tmc_blank_item_row()]
    return [{"product_id": item.product_id or "", "name": item.name, "quantity": item.quantity, "unit": item.unit} for item in instance.items.all()] or [tmc_blank_item_row()]


def tmc_item_audit_rows(instance):
    if not instance:
        return []
    return [
        {"name": item.name, "quantity": item.quantity, "unit": item.unit}
        for item in instance.items.all()
    ]


def tmc_item_audit_text(rows):
    return "; ".join(f"{row['name']} - {row['quantity']} {row['unit']}" for row in rows)


def tmc_item_change_events(old_rows, new_rows):
    old_map = {(row["name"].casefold(), row["unit"].casefold()): row for row in old_rows}
    new_map = {(row["name"].casefold(), row["unit"].casefold()): row for row in new_rows}
    events = []
    added = [new_map[key] for key in new_map.keys() - old_map.keys()]
    removed = [old_map[key] for key in old_map.keys() - new_map.keys()]
    quantity_changed = [
        {"old": old_map[key], "new": new_map[key]}
        for key in old_map.keys() & new_map.keys()
        if old_map[key]["quantity"] != new_map[key]["quantity"]
    ]
    if added:
        events.append(("tmc_item_added", "", tmc_item_audit_text(added)))
    if removed:
        events.append(("tmc_item_removed", tmc_item_audit_text(removed), ""))
    if quantity_changed:
        events.append((
            "tmc_item_quantity_changed",
            tmc_item_audit_text([item["old"] for item in quantity_changed]),
            tmc_item_audit_text([item["new"] for item in quantity_changed]),
        ))
    return events


def write_tmc_item_audit_events(obj, old_rows, new_rows, request):
    for event, old_value, new_value in tmc_item_change_events(old_rows, new_rows):
        write_audit(
            AuditLog.Action.UPDATE,
            obj,
            old_values={"items": old_value},
            new_values={"audit_event": event, "items": new_value},
            request=request,
        )


def tmc_snapshot(instance):
    data = serialize_instance(instance)
    data["items"] = "; ".join(str(item) for item in instance.items.all())
    return data


def tmc_record_form(request, organ, table, instance=None):
    if not can_write(request.user, organ, table["department"]):
        raise Http404
    old_values = tmc_snapshot(instance) if instance else None
    old_status = instance.status if instance else None
    old_item_rows = tmc_item_audit_rows(instance) if instance else []
    form = TmcRequestForm(request.POST or None, instance=instance)
    item_rows = tmc_item_rows_from_instance(instance)
    item_errors = []
    selected_photo_ids = request.POST.getlist("attached_photos") if request.method == "POST" else selected_request_photo_ids(instance)
    if request.method == "POST":
        item_rows, item_errors = tmc_item_rows_from_request(request)
        if form.is_valid() and not item_errors:
            is_create = instance is None
            with transaction.atomic():
                obj = form.save(commit=False)
                obj.territorial_organ = organ
                if not obj.pk:
                    obj.created_by = request.user
                obj.updated_by = request.user
                obj.save()
                obj.items.all().delete()
                new_item_rows = []
                for row in item_rows:
                    product, product_created = get_or_create_tmc_product(row["name"], row["unit"], row.get("product_id"))
                    if product_created:
                        write_audit(
                            AuditLog.Action.CREATE,
                            product,
                            old_values=None,
                            new_values={"audit_event": "tmc_product_created", **serialize_instance(product)},
                            request=request,
                        )
                    obj.items.create(product=product, name=product.name, quantity=row["quantity"], unit=row["unit"])
                    new_item_rows.append({"name": product.name, "quantity": row["quantity"], "unit": row["unit"]})
                photo_changes = sync_request_photos(obj, selected_photo_ids, request.user)
                if is_create or old_status != obj.status:
                    create_status_history(
                        obj=obj,
                        old_status=None if is_create else old_status,
                        new_status=obj.status,
                        completed_at=obj.due_date if obj.status == NeedStatus.DONE else None,
                        changed_by=request.user,
                        note="Создание заявки" if is_create else "Изменение статуса",
                    )
                if not is_create and old_status != obj.status:
                    write_status_change_audit_event(obj, old_status, obj.status, request)
                if not is_create:
                    write_tmc_item_audit_events(obj, old_item_rows, new_item_rows, request)
                write_request_photo_audit_events(obj, photo_changes, request)
            write_audit(AuditLog.Action.UPDATE if instance else AuditLog.Action.CREATE, obj, old_values=old_values, new_values=tmc_snapshot(obj), request=request)
            response = table_data(request, organ.pk, table["key"])
            response["HX-Trigger"] = htmx_triggers("Заявка сохранена.")
            return response
    context = {"form": form, "organ": organ, "table": table, "instance": instance, "item_rows": item_rows, "item_errors": item_errors, "show_request_photo_picker": True}
    context.update(request_photo_form_context(request, organ, selected_photo_ids))
    response = render(request, "partials/tmc_request_form.html", context)
    if request.method == "POST":
        response["HX-Retarget"] = "#modal-content"
    return response


@login_required
def tmc_product_suggest(request):
    suggestions = tmc_product_suggestions(request.GET.get("q", ""))
    return JsonResponse(
        {
            "results": [
                {"id": product.pk, "name": product.name, "unit": product.unit}
                for product in suggestions
            ]
        }
    )


@login_required
def status_history(request, organ_id, table_key, pk):
    if table_key not in STATUS_HISTORY_TABLES:
        raise Http404
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    model = REQUEST_TABLE_CONFIG[table_key]["model"]
    obj = get_object_or_404(model, pk=pk, territorial_organ=organ, is_deleted=False)
    if not can_view(request.user, organ):
        raise Http404
    return render(
        request,
        "partials/status_history.html",
        {
            "organ": organ,
            "object": obj,
            "history": status_history_queryset(obj),
            "completed_label": REQUEST_TABLE_CONFIG[table_key]["completed_label"],
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def request_photos(request, organ_id, table_key, pk):
    if table_key not in REQUEST_PHOTO_TABLES:
        raise Http404
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    model = REQUEST_TABLE_CONFIG[table_key]["model"]
    obj = get_object_or_404(model, pk=pk, territorial_organ=organ, is_deleted=False)
    if not can_view(request.user, organ):
        raise Http404
    department_slug = TABLE_BY_KEY[table_key]["department"]
    if request.method == "POST":
        if not can_write(request.user, organ, department_slug):
            raise Http404
        photo_changes = sync_request_photos(obj, request.POST.getlist("attached_photos"), request.user)
        write_request_photo_audit_events(obj, photo_changes, request)
    content_type = request_content_type_for_object(obj)
    links = (
        RequestPhotoLink.objects.select_related("photo", "photo__folder", "photo__created_by")
        .filter(territorial_organ=organ, content_type=content_type, object_id=obj.pk, photo__is_deleted=False)
        .filter(Q(photo__folder__isnull=True) | Q(photo__folder__is_deleted=False))
        .order_by("-created_at", "-id")
    )
    selected_ids = [link.photo_id for link in links]
    context = {
        "organ": organ,
        "object": obj,
        "table_key": table_key,
        "links": links,
        "can_write": can_write(request.user, organ, department_slug),
    }
    context.update(request_photo_form_context(request, organ, selected_ids))
    response = render(request, "partials/request_photos.html", context)
    if request.method == "POST":
        response["HX-Trigger"] = json.dumps({"toast": {"message": "Связанные фотографии обновлены.", "level": "success"}, "requestPhotosChanged": True})
    return response


@login_required
def request_photos_download(request, organ_id, table_key, pk):
    if table_key not in REQUEST_PHOTO_TABLES:
        raise Http404
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    model = REQUEST_TABLE_CONFIG[table_key]["model"]
    obj = get_object_or_404(model, pk=pk, territorial_organ=organ, is_deleted=False)
    if not can_view(request.user, organ):
        raise Http404
    content_type = request_content_type_for_object(obj)
    links = (
        RequestPhotoLink.objects.select_related("photo")
        .filter(territorial_organ=organ, content_type=content_type, object_id=obj.pk, photo__is_deleted=False)
        .filter(Q(photo__folder__isnull=True) | Q(photo__folder__is_deleted=False))
        .order_by("photo__created_at", "photo_id")
    )
    if not links.exists():
        raise Http404

    filename = safe_download_name(f"{obj}-photos.zip", f"request-{obj.pk}-photos.zip")
    return download_ready_response(request, photos_zip_response((link.photo for link in links), filename))


@login_required
def request_photo_picker(request, organ_id):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    context = {"organ": organ}
    context.update(request_photo_picker_context(request, organ, request.GET.getlist("attached_photos")))
    return render(request, "partials/request_photo_picker_results.html", context)


def tmc_xlsx_response(qs, organ, filename, is_multi_organ=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки ТМЦ"

    organ_offset = 1 if is_multi_organ else 0
    need_start = 1 + organ_offset
    need_end = 2 + organ_offset
    request_start = 3 + organ_offset
    request_end = 5 + organ_offset
    comment_column = 6 + organ_offset
    max_column = 6 + organ_offset

    if is_multi_organ:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(row=1, column=1, value="Территориальный орган")
    ws.merge_cells(start_row=1, start_column=need_start, end_row=1, end_column=need_end)
    ws.merge_cells(start_row=1, start_column=request_start, end_row=1, end_column=request_end)
    ws.merge_cells(start_row=1, start_column=comment_column, end_row=2, end_column=comment_column)
    ws.cell(row=1, column=need_start, value="Сведения о потребности ТМЦ")
    ws.cell(row=1, column=request_start, value="Заявка")
    ws.cell(row=1, column=comment_column, value="Описание")
    headers = ["Наименование", "Количество", "Номер", "Дата", "Исполнение заявки", ""]
    for column, value in enumerate(headers, start=need_start):
        if value:
            ws.cell(row=2, column=column, value=value)

    widths = [34, 16, 18, 14, 22, 34]
    if is_multi_organ:
        widths.insert(0, 34)
    widths = {get_column_letter(index): width for index, width in enumerate(widths, start=1)}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    subheader_fill = PatternFill("solid", fgColor="E5F1FA")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(1, 3):
        for column in range(1, max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = header_fill if row == 1 else subheader_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = Border(
                left=thin,
                right=block if column in {need_end, request_end, comment_column} else thin,
                top=thin,
                bottom=header_bottom if row == 2 else thin,
            )

    current_row = 3
    requests = list(qs)
    for request_index, obj in enumerate(requests):
        items = list(obj.items.all()) or [None]
        start_row = current_row
        end_row = current_row + len(items) - 1

        for item in items:
            ws.cell(row=current_row, column=need_start, value=item.name if item else "-")
            ws.cell(row=current_row, column=need_start + 1, value=f"{item.quantity} {item.unit}" if item else "-")
            current_row += 1

        if is_multi_organ:
            ws.cell(row=start_row, column=1, value=obj.territorial_organ.name)
        ws.cell(row=start_row, column=request_start, value=obj.request_number)
        ws.cell(row=start_row, column=request_start + 1, value=obj.request_date.strftime("%d.%m.%Y"))
        ws.cell(row=start_row, column=request_start + 2, value=obj.get_status_display())
        ws.cell(row=start_row, column=comment_column, value=obj.comment)

        if end_row > start_row:
            if is_multi_organ:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            for column in range(request_start, comment_column + 1):
                ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)

        is_last_request = request_index == len(requests) - 1
        for row in range(start_row, end_row + 1):
            for column in range(1, max_column + 1):
                cell = ws.cell(row=row, column=column)
                center_columns = {request_start, request_start + 1, request_start + 2}
                if is_multi_organ:
                    center_columns.add(1)
                cell.alignment = center_alignment if column in center_columns else body_alignment
                cell.border = Border(
                    left=thin,
                    right=block if column in {need_end, request_end, comment_column} else thin,
                    top=thin,
                    bottom=thin if is_last_request else block,
                )

    if current_row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(max_column)}{current_row - 1}"

    return workbook_file_response(wb, filename)


def tmc_grouped_export_headers(is_multi_organ):
    headers = ["Наименование ТМЦ", "Заявок"]
    if is_multi_organ:
        headers.append("Территориальных органов")
    headers.extend(["Общее количество", "Единица измерения"])
    return headers


def tmc_grouped_export_row(row, is_multi_organ):
    values = [row.get("product__name") or row.get("name") or "-", row.get("request_count") or 0]
    if is_multi_organ:
        values.append(row.get("organ_count") or 0)
    values.extend([row.get("total_quantity") or 0, row.get("unit") or ""])
    return values


def tmc_organ_grouped_export_headers():
    return ["Территориальный орган", "Заявок", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"]


def tmc_organ_grouped_export_row(row):
    return [
        row.get("request__territorial_organ__name") or "-",
        row.get("request_count") or 0,
        row.get("position_count") or 0,
        row.get("total_quantity") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def tmc_date_grouped_export_headers():
    return ["Дата", "Заявок", "Территориальных органов", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"]


def tmc_date_grouped_export_row(row):
    date = row.get("request__request_date") or row.get("request_date")
    return [
        date.strftime("%d.%m.%Y") if date else "-",
        row.get("request_count") or 0,
        row.get("organ_count") or 0,
        row.get("position_count") or 0,
        row.get("total_quantity") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def request_organ_grouped_export_headers():
    return ["Территориальный орган", "Заявок", "В работе", "Исполнено", "Отклонено"]


def request_organ_grouped_export_row(row):
    return [
        row.get("territorial_organ__name") or row.get("request__territorial_organ__name") or "-",
        row.get("request_count") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def request_date_grouped_export_headers():
    return ["Дата", "Заявок", "Территориальных органов", "В работе", "Исполнено", "Отклонено"]


def request_date_grouped_export_row(row):
    date = row.get("request_date") or row.get("request__request_date")
    return [
        date.strftime("%d.%m.%Y") if date else "-",
        row.get("request_count") or 0,
        row.get("organ_count") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def grouped_export_headers(group_mode, is_tmc=False, is_multi_organ=False):
    if group_mode == "products":
        return tmc_grouped_export_headers(is_multi_organ)
    if group_mode == "organs":
        return tmc_organ_grouped_export_headers() if is_tmc else request_organ_grouped_export_headers()
    if group_mode == "dates":
        return tmc_date_grouped_export_headers() if is_tmc else request_date_grouped_export_headers()
    return []


def grouped_export_row(row, group_mode, is_tmc=False, is_multi_organ=False):
    if group_mode == "products":
        return tmc_grouped_export_row(row, is_multi_organ)
    if group_mode == "organs":
        return tmc_organ_grouped_export_row(row) if is_tmc else request_organ_grouped_export_row(row)
    if group_mode == "dates":
        return tmc_date_grouped_export_row(row) if is_tmc else request_date_grouped_export_row(row)
    return []


def tmc_grouped_xlsx_response(rows, is_multi_organ, filename, group_mode="products"):
    wb = Workbook()
    ws = wb.active
    ws.title = "ТМЦ"
    headers = grouped_export_headers(group_mode, is_tmc=True, is_multi_organ=is_multi_organ)
    ws.append(headers)

    if group_mode == "organs":
        widths = [42, 14, 16, 18, 14, 14, 14]
    elif group_mode == "dates":
        widths = [16, 14, 24, 16, 18, 14, 14, 14]
    else:
        widths = [36, 14, 24, 18, 18] if is_multi_organ else [36, 14, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    last_column = len(headers)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=header_bottom)

    for row_index, row in enumerate(rows, start=2):
        row_values = grouped_export_row(row, group_mode, is_tmc=True, is_multi_organ=is_multi_organ)
        for column, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = body_alignment if column == 1 else center_alignment
            cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=thin)

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


def request_grouped_xlsx_response(rows, table, filename, group_mode):
    wb = Workbook()
    ws = wb.active
    ws.title = table["title"][:31]
    headers = grouped_export_headers(group_mode, is_tmc=False)
    ws.append(headers)

    widths = [42, 14, 14, 14, 14] if group_mode == "organs" else [16, 14, 24, 14, 14, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    last_column = len(headers)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=header_bottom)

    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(grouped_export_row(row, group_mode, is_tmc=False), start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = body_alignment if column == 1 else center_alignment
            cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=thin)

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


def styled_xlsx_response(qs, table, fields, filename, widths=None, center_columns=None):
    wb = Workbook()
    ws = wb.active
    ws.title = table["title"][:31]
    headers = table_header_labels(fields)
    ws.append(headers)

    widths = widths or {}
    for index, field in enumerate(fields, start=1):
        column = ws.cell(row=1, column=index).column_letter
        ws.column_dimensions[column].width = widths.get(field.name, 18)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_columns = center_columns or set()
    last_column = len(fields)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=header_bottom)

    for row_index, obj in enumerate(qs, start=2):
        for column, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_index, column=column, value=export_cell_value(obj, field))
            cell.alignment = center_alignment if field.name in center_columns else body_alignment
            cell.border = Border(left=thin, right=block if column == last_column else thin, top=thin, bottom=thin)

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


@login_required
@require_http_methods(["GET", "POST"])
def record_form(request, organ_id, table_key, pk=None):
    table = TABLE_BY_KEY[table_key]
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_write(request.user, organ, table["department"]):
        raise Http404
    instance = get_object_or_404(table["model"], pk=pk, territorial_organ=organ) if pk else None
    if table_key == "tmc-requests":
        return tmc_record_form(request, organ, table, instance)
    Form = form_for_table(table_key)
    old_values = serialize_instance(instance) if instance else None
    old_status = instance.status if instance and table_key in STATUS_HISTORY_TABLES else None
    form = Form(request.POST or None, instance=instance)
    selected_photo_ids = request.POST.getlist("attached_photos") if request.method == "POST" else selected_request_photo_ids(instance)
    if request.method == "POST" and form.is_valid():
        is_create = instance is None
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.territorial_organ = organ
            completion_field = completed_date_field(table_key)
            if table_key in STATUS_HISTORY_TABLES and obj.status == NeedStatus.DONE and not getattr(obj, completion_field):
                setattr(obj, completion_field, timezone.localdate())
            if not obj.pk:
                obj.created_by = request.user
            obj.updated_by = request.user
            obj.save()
            photo_changes = {"added": set(), "removed": set()}
            if table_key in REQUEST_PHOTO_TABLES:
                photo_changes = sync_request_photos(obj, selected_photo_ids, request.user)
            if table_key in STATUS_HISTORY_TABLES and (is_create or old_status != obj.status):
                create_status_history(
                    obj=obj,
                    old_status=None if is_create else old_status,
                    new_status=obj.status,
                    completed_at=getattr(obj, completion_field) if obj.status == NeedStatus.DONE else None,
                    changed_by=request.user,
                    note="Создание заявки" if is_create else "Изменение статуса",
                )
            if table_key in STATUS_HISTORY_TABLES and not is_create and old_status != obj.status:
                write_status_change_audit_event(obj, old_status, obj.status, request)
            if table_key in REQUEST_PHOTO_TABLES:
                write_request_photo_audit_events(obj, photo_changes, request)
        write_audit(AuditLog.Action.UPDATE if instance else AuditLog.Action.CREATE, obj, old_values=old_values, new_values=serialize_instance(obj), request=request)
        response = table_data(request, organ.pk, table_key)
        response["HX-Trigger"] = htmx_triggers("Запись сохранена.")
        return response
    context = {"form": form, "organ": organ, "table": table, "instance": instance, "show_request_photo_picker": table_key in REQUEST_PHOTO_TABLES}
    if table_key in REQUEST_PHOTO_TABLES:
        context.update(request_photo_form_context(request, organ, selected_photo_ids))
    response = render(request, "partials/record_form.html", context)
    if request.method == "POST":
        response["HX-Retarget"] = "#modal-content"
    return response


@login_required
@require_http_methods(["GET", "POST"])
def record_delete(request, organ_id, table_key, pk):
    table = TABLE_BY_KEY[table_key]
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    obj = get_object_or_404(table["model"], pk=pk, territorial_organ=organ)
    if not can_write(request.user, organ, table["department"]):
        raise Http404
    if request.method == "POST":
        old_values = serialize_instance(obj)
        obj.is_deleted = True
        obj.updated_by = request.user
        obj.save(update_fields=["is_deleted", "updated_by", "updated_at"])
        write_audit(AuditLog.Action.DELETE, obj, old_values=old_values, new_values=serialize_instance(obj), request=request)
        response = table_data(request, organ.pk, table_key)
        response["HX-Trigger"] = htmx_triggers("Запись удалена.")
        return response
    return render(request, "partials/confirm_delete.html", {"object": obj, "organ": organ, "table": table})


@login_required
def export_table(request, organ_id, table_key, fmt):
    table = TABLE_BY_KEY[table_key]
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id)
    if not can_view(request.user, organ):
        raise Http404
    selected_organs = selected_organs_from_request(request, organ)
    qs = filtered_queryset(request, table, selected_organs)
    fields = display_fields(table)
    filename = f"{table_key}-{organ.pk}.{fmt}"
    is_multi_organ = len(selected_organs) > 1
    is_request_table = table_key in REQUEST_TABLE_CONFIG
    is_fire_extinguisher_table = table_key == "fire-extinguishers"
    if table_key in STATE_SNAPSHOT_TABLES:
        qs = state_snapshot_queryset(request, table_key, qs)
    if is_fire_extinguisher_table:
        qs = fire_extinguisher_filtered_queryset(request, qs)
    current_group_mode = request_group_mode(request, table_key, is_multi_organ) if is_request_table else "requests"
    if current_group_mode in {"products", "organs", "dates"}:
        is_tmc = table_key == "tmc-requests"
        if current_group_mode == "organs":
            rows = list(tmc_organ_grouped_rows(qs) if is_tmc else request_organ_grouped_rows(qs))
        elif current_group_mode == "dates":
            rows = list(tmc_date_grouped_rows(qs) if is_tmc else request_date_grouped_rows(qs))
        else:
            rows = list(tmc_grouped_rows(qs))
        if fmt == "csv":
            csv_rows = [grouped_export_headers(current_group_mode, is_tmc=is_tmc, is_multi_organ=is_multi_organ)]
            csv_rows.extend(grouped_export_row(row, current_group_mode, is_tmc=is_tmc, is_multi_organ=is_multi_organ) for row in rows)
            return download_ready_response(request, csv_file_response(filename, csv_rows))
        if fmt == "xlsx":
            if is_tmc:
                return download_ready_response(request, tmc_grouped_xlsx_response(rows, is_multi_organ, filename, current_group_mode))
            return download_ready_response(request, request_grouped_xlsx_response(rows, table, filename, current_group_mode))
    if fmt == "csv":
        csv_rows = [table_header_labels(fields)]
        for obj in qs:
            csv_rows.append([getattr(obj, f"get_{f.name}_display", lambda: getattr(obj, f.name))() for f in fields])
        return download_ready_response(request, csv_file_response(filename, csv_rows))
    if fmt == "xlsx":
        if table_key == "tmc-requests":
            return download_ready_response(request, tmc_xlsx_response(qs, organ, filename, len(selected_organs) > 1))
        if table_key in XLSX_EXPORT_CONFIG:
            return download_ready_response(request, styled_xlsx_response(qs, table, fields, filename, **XLSX_EXPORT_CONFIG[table_key]))
        wb = Workbook()
        ws = wb.active
        ws.title = table["title"][:31]
        ws.append(table_header_labels(fields))
        for obj in qs:
            ws.append([str(getattr(obj, f"get_{f.name}_display", lambda: getattr(obj, f.name))()) for f in fields])
        return download_ready_response(request, workbook_file_response(wb, filename))
    raise Http404


@login_required
def photos(request, organ_id):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    return render_photos(request, organ)


def folder_path(folder):
    path = []
    while folder:
        if folder.is_deleted:
            return []
        path.append(folder)
        folder = folder.parent
    return list(reversed(path))


def folder_path_from_map(folder, folders_by_id):
    path = []
    while folder:
        if folder.is_deleted:
            return []
        path.append(folder)
        folder = folders_by_id.get(folder.parent_id)
    return list(reversed(path))


def add_folder_content_counts(organ, path):
    folder_ids = [folder.pk for folder in path]
    if not folder_ids:
        return path
    photo_counts = dict(
        organ.photos.filter(is_deleted=False, folder_id__in=folder_ids)
        .values("folder_id")
        .annotate(count=Count("id"))
        .values_list("folder_id", "count")
    )
    child_counts = dict(
        organ.photo_folders.filter(is_deleted=False, parent_id__in=folder_ids)
        .values("parent_id")
        .annotate(count=Count("id"))
        .values_list("parent_id", "count")
    )
    for folder in path:
        folder.breadcrumb_photo_count = photo_counts.get(folder.pk, 0)
        folder.breadcrumb_folder_count = child_counts.get(folder.pk, 0)
    return path


def add_photo_asset_permissions(user, organ, folders, photos):
    for folder in folders:
        folder.can_manage = can_manage_photo_asset(user, organ, folder)
    for photo in photos:
        photo.can_manage = can_manage_photo_asset(user, organ, photo)


def manageable_photo_folders_queryset(user, organ):
    folders = list(organ.photo_folders.select_related("parent", "created_by", "created_department").filter(is_deleted=False))
    folder_ids = [folder.pk for folder in folders if can_manage_photo_asset(user, organ, folder)]
    return organ.photo_folders.select_related("parent").filter(pk__in=folder_ids, is_deleted=False)


def can_upload_to_photo_folder(user, organ, folder):
    return can_write(user, organ) and (folder is None or can_manage_photo_asset(user, organ, folder))


def assign_photo_asset_author(obj, user):
    obj.created_by = user
    obj.updated_by = user
    obj.created_department = user_primary_department(user)


def folder_tree_is_manageable(user, organ, folder):
    folder_ids = photo_folder_descendant_ids(folder)
    folders = organ.photo_folders.filter(pk__in=folder_ids, is_deleted=False)
    photos = organ.photos.filter(folder_id__in=folder_ids, is_deleted=False)
    return all(can_manage_photo_asset(user, organ, item) for item in folders) and all(can_manage_photo_asset(user, organ, item) for item in photos)


def render_photos(request, organ, folder_id_override=None):
    query = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "newest")
    item_order = "photos" if request.GET.get("order") == "photos" else "folders"
    folder_id = str(folder_id_override) if folder_id_override is not None else request.GET.get("folder", "").strip()
    selected_folder = None
    if folder_id:
        selected_folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=folder_id, territorial_organ=organ, is_deleted=False)
    folders = organ.photo_folders.select_related("created_by", "created_department").filter(parent=selected_folder, is_deleted=False).annotate(
        photo_count=Count("photos", filter=Q(photos__is_deleted=False)),
        child_count=Count("children", filter=Q(children__is_deleted=False), distinct=True),
    )
    folders = folders.order_by("created_at", "pk") if sort == "oldest" else folders.order_by("-created_at", "-pk")
    if query:
        query_normalized = query.casefold()
        folders = [folder for folder in folders if query_normalized in folder.name.casefold()]
    qs = organ.photos.select_related("created_by", "created_department", "folder").filter(is_deleted=False).filter(Q(folder__isnull=True) | Q(folder__is_deleted=False))
    if folder_id:
        qs = qs.filter(folder_id=folder_id)
    else:
        qs = qs.filter(folder__isnull=True)
    if query:
        qs = [photo for photo in qs if photo_matches_query(photo, query)]
        qs = sorted(qs, key=lambda photo: (photo.created_at, photo.pk), reverse=sort != "oldest")
    else:
        qs = qs.order_by("created_at", "pk") if sort == "oldest" else qs.order_by("-created_at", "-pk")
    paginator = Paginator(qs, 24)
    page = paginator.get_page(request.GET.get("page"))
    page_links = paginator.get_elided_page_range(page.number, on_each_side=1, on_ends=1)
    folders_by_id = {folder.pk: folder for folder in organ.photo_folders.filter(is_deleted=False)}
    for photo in page.object_list:
        photo.folder_path = folder_path_from_map(photo.folder, folders_by_id) if photo.folder else []
    add_photo_asset_permissions(request.user, organ, folders, page.object_list)
    querystring = request.GET.copy()
    querystring.pop("page", None)
    folder_path_items = add_folder_content_counts(organ, folder_path(selected_folder))
    root_photo_count = organ.photos.filter(is_deleted=False, folder__isnull=True).count()
    root_folder_count = organ.photo_folders.filter(is_deleted=False, parent__isnull=True).count()
    total_photo_count = organ.photos.filter(is_deleted=False).filter(Q(folder__isnull=True) | Q(folder__is_deleted=False)).count()
    total_folder_count = len(folders_by_id)
    can_upload_photos = can_upload_to_photo_folder(request.user, organ, selected_folder)
    return render(
        request,
        "partials/photos.html",
        {
            "organ": organ,
            "photos": page.object_list,
            "photo_page": page,
            "photo_page_links": page_links,
            "photo_querystring": querystring.urlencode(),
            "folders": folders,
            "photo_folder_count": len(folders),
            "photo_summary_count": page.paginator.count,
            "photo_summary_folder_count": len(folders),
            "photo_total_count": total_photo_count,
            "photo_total_folder_count": total_folder_count,
            "photo_root_count": root_photo_count,
            "photo_root_folder_count": root_folder_count,
            "selected_folder": selected_folder,
            "folder_path": folder_path_items,
            "photo_folder": folder_id,
            "can_write": can_write(request.user, organ),
            "can_upload_photos": can_upload_photos,
            "photo_query": query,
            "photo_sort": sort,
            "photo_item_order": item_order,
        },
    )


def safe_download_name(value, fallback):
    name = "".join(char if char.isalnum() or char in "._- " else "_" for char in value).strip()
    return name or fallback


def photo_download_name(photo):
    return safe_download_name(photo.original_filename or Path(photo.image.name).name, f"photo-{photo.pk}")


def unique_archive_name(relative_name, photo_pk, used_names):
    path = PurePosixPath(relative_name)
    parent = "" if str(path.parent) == "." else f"{path.parent}/"
    source_name = path.name
    stem = Path(source_name).stem
    suffix = Path(source_name).suffix
    archive_name = relative_name
    counter = 1
    while archive_name in used_names:
        extra = f"-{photo_pk}" if counter == 1 else f"-{photo_pk}-{counter}"
        archive_name = f"{parent}{stem}{extra}{suffix}"
        counter += 1
    used_names.add(archive_name)
    return archive_name


def photos_zip_response(photos, filename, archive_path_builder=None):
    temp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
    path = temp.name
    temp.close()
    used_names = set()
    try:
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            for index, photo in enumerate(photos, start=1):
                if not photo.image:
                    continue
                source_name = photo_download_name(photo)
                relative_name = archive_path_builder(photo, source_name) if archive_path_builder else f"{index:03d}-{source_name}"
                archive_name = unique_archive_name(relative_name, photo.pk, used_names)
                try:
                    with photo.image.open("rb") as file_handle:
                        zip_file.writestr(archive_name, file_handle.read())
                except FileNotFoundError:
                    continue
    except Exception:
        try:
            os.remove(path)
        except FileNotFoundError:
            pass
        raise
    return temporary_download_response(path, filename, "application/zip")


@login_required
def photo_download(request, organ_id, pk):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    photo = get_object_or_404(
        TerritorialOrganPhoto.objects.filter(Q(folder__isnull=True) | Q(folder__is_deleted=False)),
        pk=pk,
        territorial_organ=organ,
        is_deleted=False,
    )
    if not photo.image:
        raise Http404
    try:
        file_handle = photo.image.open("rb")
    except FileNotFoundError:
        raise Http404
    filename = photo_download_name(photo)
    content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    return FileResponse(file_handle, as_attachment=True, filename=filename, content_type=content_type)


@login_required
def photos_download_all(request, organ_id):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    photos_qs = organ.photos.filter(is_deleted=False).filter(Q(folder__isnull=True) | Q(folder__is_deleted=False)).order_by("created_at")
    if not photos_qs.exists():
        raise Http404
    filename = safe_download_name(f"{organ.name}-photos.zip", f"organ-{organ.pk}-photos.zip")
    return download_ready_response(request, photos_zip_response(photos_qs, filename))


@login_required
def photo_folder_download(request, organ_id, pk):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_view(request.user, organ):
        raise Http404
    folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=pk, territorial_organ=organ, is_deleted=False)
    folder_ids = photo_folder_descendant_ids(folder)
    photos_qs = (
        organ.photos.select_related("folder")
        .filter(is_deleted=False, folder_id__in=folder_ids)
        .filter(folder__is_deleted=False)
        .order_by("folder_id", "created_at", "pk")
    )
    if not photos_qs.exists():
        raise Http404
    folders_by_id = {item.pk: item for item in organ.photo_folders.filter(pk__in=folder_ids, is_deleted=False)}
    root_index = folder_path_from_map(folder, folders_by_id)

    def archive_path(photo, source_name):
        current_path = folder_path_from_map(photo.folder, folders_by_id)
        nested_path = current_path[len(root_index):]
        parts = [safe_download_name(item.name, f"folder-{item.pk}") for item in nested_path]
        return "/".join([*parts, source_name]) if parts else source_name

    filename = safe_download_name(f"{folder.name}-photos.zip", f"folder-{folder.pk}-photos.zip")
    return download_ready_response(request, photos_zip_response(photos_qs, filename, archive_path))


@login_required
@require_http_methods(["GET", "POST"])
def photo_form(request, organ_id, pk=None):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_write(request.user, organ):
        raise Http404
    photo = get_object_or_404(TerritorialOrganPhoto, pk=pk, territorial_organ=organ) if pk else None
    if photo and not can_manage_photo_asset(request.user, organ, photo):
        raise Http404
    old_values = serialize_instance(photo) if photo else None
    folder_queryset = manageable_photo_folders_queryset(request.user, organ)
    form = TerritorialOrganPhotoForm(request.POST or None, request.FILES or None, instance=photo, organ=organ, folder_queryset=folder_queryset)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.territorial_organ = organ
        if photo and request.FILES.get("image"):
            obj.created_at = timezone.now()
        if not obj.pk:
            assign_photo_asset_author(obj, request.user)
        else:
            obj.updated_by = request.user
        obj.save()
        write_audit(AuditLog.Action.UPDATE if photo else AuditLog.Action.CREATE, obj, old_values=old_values, new_values=serialize_instance(obj), request=request)
        response = render_photos(request, organ, obj.folder_id or "")
        response["HX-Trigger"] = htmx_triggers("Фотография сохранена.")
        return response
    return render(request, "partials/photo_form.html", {"form": form, "organ": organ, "photo": photo})


@login_required
@require_http_methods(["GET", "POST"])
def photo_folder_form(request, organ_id, pk=None):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_write(request.user, organ):
        raise Http404
    folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=pk, territorial_organ=organ, is_deleted=False) if pk else None
    if folder and not can_manage_photo_asset(request.user, organ, folder):
        raise Http404
    old_values = serialize_instance(folder) if folder else None
    parent_id = request.POST.get("parent") if request.method == "POST" else request.GET.get("folder")
    current_folder = folder.parent if folder else None
    if parent_id and not folder:
        current_folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=parent_id, territorial_organ=organ, is_deleted=False)
    if not can_upload_to_photo_folder(request.user, organ, current_folder):
        raise Http404
    folder_queryset = manageable_photo_folders_queryset(request.user, organ)
    form = TerritorialOrganPhotoFolderForm(request.POST or None, instance=folder, organ=organ, parent=current_folder, folder_queryset=folder_queryset)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.territorial_organ = organ
        if not folder:
            obj.parent = current_folder
            assign_photo_asset_author(obj, request.user)
        elif "parent" not in request.POST:
            obj.parent = current_folder
            obj.updated_by = request.user
        else:
            obj.updated_by = request.user
        obj.save()
        write_audit(AuditLog.Action.UPDATE if folder else AuditLog.Action.CREATE, obj, old_values=old_values, new_values=serialize_instance(obj), request=request)
        response = render_photos(request, organ, obj.parent_id or "")
        response["HX-Trigger"] = htmx_triggers("Папка переименована." if folder else "Папка создана.")
        return response
    return render(request, "partials/photo_folder_form.html", {"form": form, "organ": organ, "folder": folder, "current_folder": current_folder})


def photo_folder_descendant_ids(folder):
    folder_ids = [folder.pk]
    pending = [folder.pk]
    while pending:
        child_ids = list(TerritorialOrganPhotoFolder.objects.filter(parent_id__in=pending, is_deleted=False).values_list("pk", flat=True))
        folder_ids.extend(child_ids)
        pending = child_ids
    return folder_ids


@login_required
@require_http_methods(["GET", "POST"])
def photo_folder_delete(request, organ_id, pk):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=pk, territorial_organ=organ, is_deleted=False)
    if not can_manage_photo_asset(request.user, organ, folder) or not folder_tree_is_manageable(request.user, organ, folder):
        raise Http404
    parent = folder.parent
    if request.method == "POST":
        with transaction.atomic():
            old_values = serialize_instance(folder)
            folder_ids = photo_folder_descendant_ids(folder)
            TerritorialOrganPhoto.objects.filter(territorial_organ=organ, folder_id__in=folder_ids, is_deleted=False).update(
                is_deleted=True,
                updated_by=request.user,
                updated_at=timezone.now(),
            )
            TerritorialOrganPhotoFolder.objects.filter(territorial_organ=organ, pk__in=folder_ids).update(is_deleted=True, updated_by=request.user, updated_at=timezone.now())
            write_audit(AuditLog.Action.DELETE, folder, old_values=old_values, new_values=None, request=request)
        response = render_photos(request, organ, parent.pk if parent else "")
        response["HX-Trigger"] = htmx_triggers("Папка удалена.")
        return response
    return render(request, "partials/confirm_delete.html", {"object": folder, "organ": organ, "folder_delete": True})


@login_required
@require_http_methods(["GET", "POST"])
def photo_bulk_upload(request, organ_id):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    if not can_write(request.user, organ):
        raise Http404
    current_folder = None
    if request.GET.get("folder"):
        current_folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=request.GET["folder"], territorial_organ=organ, is_deleted=False)
    if not can_upload_to_photo_folder(request.user, organ, current_folder):
        raise Http404
    if request.method == "POST":
        files = request.FILES.getlist("images")
        descriptions = request.POST.getlist("descriptions")
        folder = None
        folder_id = request.POST.get("folder")
        if folder_id and current_folder is None:
            current_folder = get_object_or_404(TerritorialOrganPhotoFolder, pk=folder_id, territorial_organ=organ, is_deleted=False)
        if not can_upload_to_photo_folder(request.user, organ, current_folder):
            raise Http404
        new_folder_name = request.POST.get("new_folder", "").strip()
        if new_folder_name:
            folder = TerritorialOrganPhotoFolder.objects.filter(territorial_organ=organ, parent=current_folder, name=new_folder_name, is_deleted=False).first()
            if folder and not can_manage_photo_asset(request.user, organ, folder):
                raise Http404
            if folder is None:
                folder = TerritorialOrganPhotoFolder(territorial_organ=organ, parent=current_folder, name=new_folder_name)
                assign_photo_asset_author(folder, request.user)
                folder.save()
        elif folder_id:
            folder = current_folder
        errors = []
        created = 0
        for index, image in enumerate(files):
            data = {"description": descriptions[index] if index < len(descriptions) else ""}
            if folder:
                data["folder"] = folder.pk
            form = TerritorialOrganPhotoForm(data, {"image": image}, organ=organ, folder_queryset=manageable_photo_folders_queryset(request.user, organ))
            if form.is_valid():
                obj = form.save(commit=False)
                obj.territorial_organ = organ
                assign_photo_asset_author(obj, request.user)
                obj.save()
                write_audit(AuditLog.Action.CREATE, obj, old_values=None, new_values=serialize_instance(obj), request=request)
                created += 1
            else:
                errors.append(f"{image.name}: {form.errors.as_text()}")
        if request.headers.get("X-Bulk-Photo-Batch") == "true":
            return JsonResponse(
                {
                    "created": created,
                    "failed": len(errors),
                    "errors": errors[:10],
                    "folder": folder.pk if folder else None,
                }
            )
        response = render_photos(request, organ, folder.pk if folder else "")
        if errors:
            response["HX-Trigger"] = htmx_triggers(f"Загружено: {created}. Не загружено: {len(errors)}.", "warning")
        else:
            response["HX-Trigger"] = htmx_triggers(f"Фотографий загружено: {created}.")
        return response
    return render(request, "partials/photo_bulk_form.html", {"organ": organ, "current_folder": current_folder})


@login_required
@require_http_methods(["GET", "POST"])
def photo_delete(request, organ_id, pk):
    organ = get_object_or_404(TerritorialOrgan, pk=organ_id, is_active=True)
    photo = get_object_or_404(TerritorialOrganPhoto, pk=pk, territorial_organ=organ)
    if not can_manage_photo_asset(request.user, organ, photo):
        raise Http404
    if request.method == "POST":
        old_values = serialize_instance(photo)
        photo.is_deleted = True
        photo.updated_by = request.user
        photo.save(update_fields=["is_deleted", "updated_by", "updated_at"])
        write_audit(AuditLog.Action.DELETE, photo, old_values=old_values, new_values=serialize_instance(photo), request=request)
        response = render_photos(request, organ, photo.folder_id or "")
        response["HX-Trigger"] = htmx_triggers("Фотография удалена.")
        return response
    return render(request, "partials/confirm_delete.html", {"object": photo, "organ": organ, "photo_delete": True})
