"""Query-count regression tests for the pages users hit most.

These aren't meant to force an exact "ideal" query count - just to catch a
future N+1 regression before it ships. Bounds are set with headroom over the
count measured at the time the test was written (noted per test), not a
target to shrink toward.
"""
import time

from django.db import connection
from django.test.utils import CaptureQueriesContext
from openpyxl import Workbook

from .services.exports import fast_merge_cells
from .tests_base import *


class PerformanceRegressionTests(RequestAppTestCase):
    def seed_tmc_requests(self, count=10):
        for index in range(count):
            obj = TmcRequest.objects.create(
                territorial_organ=self.organ,
                created_by=self.user,
                request_number=f"perf-{index}",
                request_date="2026-06-20",
                status="in_work",
            )
            TmcRequestItem.objects.create(request=obj, name="Item", quantity=1, unit="шт.")

    def test_dashboard_query_count_has_a_ceiling(self):
        # Measured 12 queries for 10 seeded requests at write time.
        self.seed_tmc_requests()
        self.client.login(username="operator", password="pass12345")

        # Warm request first: the user-menu trash badge computes its count
        # once per cache TTL, so steady-state page cost - what this ceiling
        # protects - is measured on the second render.
        self.client.get(reverse("dashboard"))
        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries.captured_queries), 20)

    def test_fast_merge_cells_scales_linearly_not_quadratically(self):
        # ws.merge_cells() scans every previously-merged range on the sheet
        # to check for overlaps on each call, which turns thousands of calls
        # in a loop into an effectively O(n^2) wall-clock cost - measured
        # 12.3s for 8,000 merges alone, extrapolating to well over an hour
        # for a real ~35k-row multi-item TMC export (tmc_xlsx_response merges
        # a range per multi-item request). fast_merge_cells() skips that
        # check since tmc_xlsx_response's ranges are always disjoint by
        # construction. 10,000 merges measured ~0.9s with the fix; the old
        # ws.merge_cells() path would take ~19s at this scale, so 5s of
        # headroom is a clear signal if this regresses back to O(n^2).
        wb = Workbook()
        ws = wb.active
        start = time.time()
        for index in range(10000):
            row = index * 2 + 1
            fast_merge_cells(ws, row, 1, row + 1, 1)
        elapsed = time.time() - start

        self.assertEqual(len(ws.merged_cells.ranges), 10000)
        self.assertLess(elapsed, 5, "fast_merge_cells regressed back to roughly O(n^2) scaling")

    def test_htmx_table_load_query_count_has_a_ceiling(self):
        # Measured 20 queries for 10 seeded requests at write time.
        self.seed_tmc_requests()
        self.client.login(username="operator", password="pass12345")

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries.captured_queries), 30)

    def test_photo_gallery_query_count_has_a_ceiling(self):
        # Measured 27 queries for 10 photos at write time (mild sub-linear
        # scaling observed - not flagged as a bug here, just given headroom).
        for index in range(10):
            self.create_photo(f"perf-{index}.png")
        self.client.login(username="operator", password="pass12345")

        # Warm request first: the user-menu trash badge computes its count
        # once per cache TTL, so steady-state page cost - what this ceiling
        # protects - is measured on the second render.
        self.client.get(reverse("photos", args=[self.organ.pk]))
        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("photos", args=[self.organ.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries.captured_queries), 40)

    def test_csv_export_query_count_has_a_ceiling(self):
        # Measured 10 queries for 10 seeded requests at write time.
        self.seed_tmc_requests()
        self.client.login(username="operator", password="pass12345")

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("export_table", args=[self.organ.pk, "tmc-requests", "csv"]))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries.captured_queries), 18)

    def test_tmc_xlsx_export_query_count_has_a_ceiling(self):
        # Measured 13 queries for 10 seeded requests at write time.
        self.seed_tmc_requests()
        self.client.login(username="operator", password="pass12345")

        with CaptureQueriesContext(connection) as queries:
            response = self.client.get(reverse("export_table", args=[self.organ.pk, "tmc-requests", "xlsx"]))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(len(queries.captured_queries), 20)
