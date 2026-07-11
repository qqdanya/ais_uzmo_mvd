"""Excel/CSV export helpers for request and state tables."""

from types import SimpleNamespace

from django.utils.text import capfirst
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.requests_app.services.downloads import workbook_file_response

# All styled XLSX exports below always keep their formatting regardless of
# row count - this only decides which exports are "large" enough to need a
# heavy_export_slot() (see table_exports.py/export_limits.py), so a handful
# of big exports can't tie up every gunicorn worker at once.
XLSX_WRITE_ONLY_THRESHOLD = 5000


def display_fields(table):
    field_names = table["fields"]
    fields_by_name = {field.name: field for field in table["model"]._meta.fields}
    computed_labels = {"items_summary": "наименования"}
    return [fields_by_name.get(name) or SimpleNamespace(name=name, verbose_name=computed_labels.get(name, name)) for name in field_names]


def table_header_labels(fields):
    return [capfirst(field.verbose_name) for field in fields]


def export_cell_value(obj, field):
    display = getattr(obj, f"get_{field.name}_display", None)
    value = display() if callable(display) else getattr(obj, field.name)
    if hasattr(value, "strftime") and getattr(field, "get_internal_type", lambda: "")() == "DateField":
        return value.strftime("%d.%m.%Y")
    return value


def export_objects(qs):
    return qs.iterator(chunk_size=1000) if hasattr(qs, "iterator") else qs


def export_row_count(rows):
    # list has its own .count(value) (counts occurrences of a value, not the
    # list's length), so it has to be checked before the queryset .count()
    # branch below or this misfires with a TypeError - grouped rows
    # (grouping.py's request_organ_grouped_rows/request_date_grouped_rows)
    # are plain lists, not querysets.
    if isinstance(rows, list):
        return len(rows)
    if hasattr(rows, "count") and callable(rows.count):
        return rows.count()
    if hasattr(rows, "__len__"):
        return len(rows)
    return None


def should_use_write_only(rows):
    count = export_row_count(rows)
    return count is None or count > XLSX_WRITE_ONLY_THRESHOLD


def write_only_xlsx_response(title, headers, rows, filename):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return workbook_file_response(wb, filename)


def tmc_xlsx_response(qs, organ, filename, is_multi_organ=False):
    # Callers are expected to already prefetch "items", but re-asserting it
    # here is cheap (Django no-ops a repeated prefetch) and keeps this
    # function safe against N+1 if it's ever called with a bare queryset.
    if hasattr(qs, "prefetch_related"):
        qs = qs.prefetch_related("items")

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки ТМЦ"

    organ_offset = 1 if is_multi_organ else 0
    need_start = 1 + organ_offset
    need_end = 2 + organ_offset
    request_start = 3 + organ_offset
    request_end = 5 + organ_offset
    comment_column = 6 + organ_offset
    max_column = 6 + organ_offset

    if is_multi_organ:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(row=1, column=1, value="Территориальный орган")
    ws.merge_cells(start_row=1, start_column=need_start, end_row=1, end_column=need_end)
    ws.merge_cells(start_row=1, start_column=request_start, end_row=1, end_column=request_end)
    ws.merge_cells(start_row=1, start_column=comment_column, end_row=2, end_column=comment_column)
    ws.cell(row=1, column=need_start, value="Сведения о потребности ТМЦ")
    ws.cell(row=1, column=request_start, value="Заявка")
    ws.cell(row=1, column=comment_column, value="Описание")
    headers = ["Наименование", "Количество", "Номер", "Дата", "Исполнение заявки", ""]
    for column, value in enumerate(headers, start=need_start):
        if value:
            ws.cell(row=2, column=column, value=value)

    widths = [34, 16, 18, 14, 22, 34]
    if is_multi_organ:
        widths.insert(0, 34)
    widths = {get_column_letter(index): width for index, width in enumerate(widths, start=1)}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    subheader_fill = PatternFill("solid", fgColor="E5F1FA")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Every cell's border is one of a handful of fixed combinations - build
    # each once and reuse it instead of a fresh Border() per cell, which is
    # most of the cost of writing a large styled sheet (openpyxl interns
    # style objects into a shared table, so re-passing the same instance is
    # a table lookup instead of a new entry each time).
    block_columns = {need_end, request_end, comment_column}
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border_block = Border(left=thin, right=block, top=thin, bottom=thin)
    subheader_border = Border(left=thin, right=thin, top=thin, bottom=header_bottom)
    subheader_border_block = Border(left=thin, right=block, top=thin, bottom=header_bottom)
    body_border = Border(left=thin, right=thin, top=thin, bottom=block)
    body_border_block = Border(left=thin, right=block, top=thin, bottom=block)

    for row in range(1, 3):
        for column in range(1, max_column + 1):
            cell = ws.cell(row=row, column=column)
            cell.fill = header_fill if row == 1 else subheader_fill
            cell.font = header_font
            cell.alignment = center_alignment
            is_block_column = column in block_columns
            if row == 1:
                cell.border = header_border_block if is_block_column else header_border
            else:
                cell.border = subheader_border_block if is_block_column else subheader_border

    body_center_columns = {request_start, request_start + 1, request_start + 2}
    if is_multi_organ:
        body_center_columns.add(1)

    current_row = 3
    for obj in export_objects(qs):
        items = list(obj.items.all()) or [None]
        start_row = current_row
        end_row = current_row + len(items) - 1

        for item in items:
            ws.cell(row=current_row, column=need_start, value=item.name if item else "-")
            ws.cell(row=current_row, column=need_start + 1, value=f"{item.quantity} {item.unit}" if item else "-")
            current_row += 1

        if is_multi_organ:
            ws.cell(row=start_row, column=1, value=obj.territorial_organ.name)
        ws.cell(row=start_row, column=request_start, value=obj.request_number)
        ws.cell(row=start_row, column=request_start + 1, value=obj.request_date.strftime("%d.%m.%Y"))
        ws.cell(row=start_row, column=request_start + 2, value=obj.get_status_display())
        ws.cell(row=start_row, column=comment_column, value=obj.comment)

        if end_row > start_row:
            if is_multi_organ:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            for column in range(request_start, comment_column + 1):
                ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)

        for row in range(start_row, end_row + 1):
            for column in range(1, max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.alignment = center_alignment if column in body_center_columns else body_alignment
                cell.border = body_border_block if column in block_columns else body_border

    if current_row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(max_column)}{current_row - 1}"

    return workbook_file_response(wb, filename)


def tmc_grouped_export_headers(is_multi_organ):
    headers = ["Наименование ТМЦ", "Заявок"]
    if is_multi_organ:
        headers.append("Территориальных органов")
    headers.extend(["Общее количество", "Единица измерения"])
    return headers


def tmc_grouped_export_row(row, is_multi_organ):
    values = [row.get("product__name") or row.get("name") or "-", row.get("request_count") or 0]
    if is_multi_organ:
        values.append(row.get("organ_count") or 0)
    values.extend([row.get("total_quantity") or 0, row.get("unit") or ""])
    return values


def tmc_organ_grouped_export_headers():
    return ["Территориальный орган", "Заявок", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"]


def tmc_organ_grouped_export_row(row):
    return [
        row.get("request__territorial_organ__name") or "-",
        row.get("request_count") or 0,
        row.get("position_count") or 0,
        row.get("total_quantity") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def tmc_date_grouped_export_headers():
    return ["Дата", "Заявок", "Территориальных органов", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"]


def tmc_date_grouped_export_row(row):
    date = row.get("request__request_date") or row.get("request_date")
    return [
        date.strftime("%d.%m.%Y") if date else "-",
        row.get("request_count") or 0,
        row.get("organ_count") or 0,
        row.get("position_count") or 0,
        row.get("total_quantity") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def request_organ_grouped_export_headers():
    return ["Территориальный орган", "Заявок", "В работе", "Исполнено", "Отклонено"]


def request_organ_grouped_export_row(row):
    return [
        row.get("territorial_organ__name") or row.get("request__territorial_organ__name") or "-",
        row.get("request_count") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def request_date_grouped_export_headers():
    return ["Дата", "Заявок", "Территориальных органов", "В работе", "Исполнено", "Отклонено"]


def request_date_grouped_export_row(row):
    date = row.get("request_date") or row.get("request__request_date")
    return [
        date.strftime("%d.%m.%Y") if date else "-",
        row.get("request_count") or 0,
        row.get("organ_count") or 0,
        row.get("in_work_count") or 0,
        row.get("done_count") or 0,
        row.get("rejected_count") or 0,
    ]


def grouped_export_headers(group_mode, is_tmc=False, is_multi_organ=False):
    if group_mode == "products":
        return tmc_grouped_export_headers(is_multi_organ)
    if group_mode == "organs":
        return tmc_organ_grouped_export_headers() if is_tmc else request_organ_grouped_export_headers()
    if group_mode == "dates":
        return tmc_date_grouped_export_headers() if is_tmc else request_date_grouped_export_headers()
    return []


def grouped_export_row(row, group_mode, is_tmc=False, is_multi_organ=False):
    if group_mode == "products":
        return tmc_grouped_export_row(row, is_multi_organ)
    if group_mode == "organs":
        return tmc_organ_grouped_export_row(row) if is_tmc else request_organ_grouped_export_row(row)
    if group_mode == "dates":
        return tmc_date_grouped_export_row(row) if is_tmc else request_date_grouped_export_row(row)
    return []


def tmc_grouped_xlsx_response(rows, is_multi_organ, filename, group_mode="products"):
    wb = Workbook()
    ws = wb.active
    ws.title = "ТМЦ"
    headers = grouped_export_headers(group_mode, is_tmc=True, is_multi_organ=is_multi_organ)
    ws.append(headers)

    if group_mode == "organs":
        widths = [42, 14, 16, 18, 14, 14, 14]
    elif group_mode == "dates":
        widths = [16, 14, 24, 16, 18, 14, 14, 14]
    else:
        widths = [36, 14, 24, 18, 18] if is_multi_organ else [36, 14, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    last_column = len(headers)

    # Build each border variant once and reuse it, instead of a fresh
    # Border() per cell - see the comment in tmc_xlsx_response.
    header_border = Border(left=thin, right=thin, top=thin, bottom=header_bottom)
    header_border_last = Border(left=thin, right=block, top=thin, bottom=header_bottom)
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_border_last = Border(left=thin, right=block, top=thin, bottom=thin)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = header_border_last if column == last_column else header_border

    for row_index, row in enumerate(rows, start=2):
        row_values = grouped_export_row(row, group_mode, is_tmc=True, is_multi_organ=is_multi_organ)
        for column, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = body_alignment if column == 1 else center_alignment
            cell.border = body_border_last if column == last_column else body_border

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


def request_grouped_xlsx_response(rows, table, filename, group_mode):
    wb = Workbook()
    ws = wb.active
    ws.title = table["title"][:31]
    headers = grouped_export_headers(group_mode, is_tmc=False)
    ws.append(headers)

    widths = [42, 14, 14, 14, 14] if group_mode == "organs" else [16, 14, 24, 14, 14, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    last_column = len(headers)

    header_border = Border(left=thin, right=thin, top=thin, bottom=header_bottom)
    header_border_last = Border(left=thin, right=block, top=thin, bottom=header_bottom)
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_border_last = Border(left=thin, right=block, top=thin, bottom=thin)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = header_border_last if column == last_column else header_border

    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(grouped_export_row(row, group_mode, is_tmc=False), start=1):
            cell = ws.cell(row=row_index, column=column, value=value)
            cell.alignment = body_alignment if column == 1 else center_alignment
            cell.border = body_border_last if column == last_column else body_border

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


def styled_xlsx_response(qs, table, fields, filename, widths=None, center_columns=None):
    wb = Workbook()
    ws = wb.active
    ws.title = table["title"][:31]
    headers = table_header_labels(fields)
    ws.append(headers)

    widths = widths or {}
    for index, field in enumerate(fields, start=1):
        column = ws.cell(row=1, column=index).column_letter
        ws.column_dimensions[column].width = widths.get(field.name, 18)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="C6DBE9")
    block = Side(style="medium", color="7FAED0")
    header_bottom = Side(style="medium", color="8FBFDD")
    header_fill = PatternFill("solid", fgColor="D6EAF7")
    header_font = Font(bold=True, color="0B2F5B")
    body_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_columns = center_columns or set()
    last_column = len(fields)

    header_border = Border(left=thin, right=thin, top=thin, bottom=header_bottom)
    header_border_last = Border(left=thin, right=block, top=thin, bottom=header_bottom)
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_border_last = Border(left=thin, right=block, top=thin, bottom=thin)

    for column in range(1, last_column + 1):
        cell = ws.cell(row=1, column=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = header_border_last if column == last_column else header_border

    for row_index, obj in enumerate(export_objects(qs), start=2):
        for column, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_index, column=column, value=export_cell_value(obj, field))
            cell.alignment = center_alignment if field.name in center_columns else body_alignment
            cell.border = body_border_last if column == last_column else body_border

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=last_column).column_letter}{ws.max_row}"

    return workbook_file_response(wb, filename)


def basic_xlsx_response(qs, table, fields, filename):
    rows = (
        [str(getattr(obj, f"get_{field.name}_display", lambda: getattr(obj, field.name))()) for field in fields]
        for obj in export_objects(qs)
    )
    return write_only_xlsx_response(table["title"], table_header_labels(fields), rows, filename)
