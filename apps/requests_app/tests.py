import csv
from io import BytesIO
from datetime import timedelta
from urllib.parse import quote_plus
import zipfile

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image

from apps.accounts.models import UserProfile
from apps.audit.models import AuditLog
from apps.directory.models import Department, TerritorialOrgan, TerritorialOrganPhoto, TerritorialOrganPhotoFolder
from apps.requests_app.models import (
    ACTIVE_NEED_STATUS_CHOICES,
    AntiTerrorMeasure,
    CitsiziEquipment,
    EquipmentType,
    FireAlarm,
    FireDepartmentRequest,
    FireExtinguisher,
    SecurityAlarm,
    BuildingRepairRequest,
    RequestPhotoLink,
    RequestStatusHistory,
    ServiceHousing,
    TmcProduct,
    TmcRequest,
    TmcRequestItem,
    VehicleFuelRequest,
    VehicleInventory,
    VehicleRepairRequest,
)
from apps.requests_app.registry import TABLES, TABLE_BY_KEY


class AppFlowTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user("operator", password="pass12345")
        UserProfile.objects.create(user=self.user, role=UserProfile.Role.OPERATOR)
        self.organ = TerritorialOrgan.objects.create(name="Test territorial organ", order_number=1)
        self.department = Department.objects.create(name="TMC", slug="tmc", order_number=1)

    def status_history(self, obj):
        content_type = ContentType.objects.get_for_model(obj, for_concrete_model=False)
        return RequestStatusHistory.objects.filter(content_type=content_type, object_id=obj.pk)

    def create_status_history_entry(self, obj, old_status=None, new_status="in_work"):
        return RequestStatusHistory.objects.create(
            content_type=ContentType.objects.get_for_model(obj, for_concrete_model=False),
            object_id=obj.pk,
            old_status=old_status,
            new_status=new_status,
            changed_by=self.user,
        )

    def response_bytes(self, response):
        if getattr(response, "streaming", False):
            return b"".join(response.streaming_content)
        return response.content

    def response_workbook(self, response):
        return load_workbook(BytesIO(self.response_bytes(response)))

    def test_dashboard_requires_login(self):
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 302)

    def test_dashboard_after_login(self):
        Department.objects.create(name="Transport", slug="transport", order_number=2)
        Department.objects.create(name="Unknown", slug="unknown", order_number=3)
        self.client.login(username="operator", password="pass12345")
        response = self.client.get(reverse("dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Test territorial organ")
        self.assertContains(response, "bi-box-seam")
        self.assertContains(response, "bi-truck")
        self.assertContains(response, "bi-folder2-open")

    def test_crud_creates_tmc_request_with_multiple_items_and_audit_log(self):
        self.client.login(username="operator", password="pass12345")
        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "tmc-requests"]),
            {
                "request_number": "15/TMC",
                "request_date": "2026-06-27",
                "status": "in_work",
                "comment": "Urgent",
                "item_name": ["Paper", "Keyboard"],
                "item_quantity": ["10", "3"],
                "item_unit": ["pack", "pcs"],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        request_obj = TmcRequest.objects.get(request_number="15/TMC")
        self.assertEqual(request_obj.items.count(), 2)
        self.assertTrue(request_obj.items.filter(name="Paper", quantity=10, unit="pack").exists())
        self.assertTrue(request_obj.items.filter(name="Keyboard", quantity=3, unit="pcs").exists())
        self.assertTrue(TmcProduct.objects.filter(name="Paper", normalized_name="paper").exists())
        self.assertTrue(request_obj.items.filter(name="Paper", product__name="Paper").exists())
        self.assertTrue(self.status_history(request_obj).filter(old_status__isnull=True, new_status="in_work", changed_by=self.user).exists())
        self.assertTrue(AuditLog.objects.filter(action=AuditLog.Action.CREATE, model_name="TmcRequest").exists())
        self.assertNotContains(response, "request-photo-count")

    def test_tmc_request_uses_selected_product_from_suggestion(self):
        product = TmcProduct.objects.create(name="Стол компьютерный", unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "tmc-requests"]),
            {
                "request_number": "16/TMC",
                "request_date": "2026-06-27",
                "status": "in_work",
                "comment": "Selected product",
                "item_product": [str(product.pk)],
                "item_name": ["компьютерный стол"],
                "item_quantity": ["2"],
                "item_unit": ["шт."],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        item = TmcRequest.objects.get(request_number="16/TMC").items.get()
        self.assertEqual(item.product, product)
        self.assertEqual(item.name, "Стол компьютерный")
        self.assertEqual(TmcProduct.objects.count(), 1)

    def test_tmc_request_creates_new_product_when_suggestion_not_selected(self):
        TmcProduct.objects.create(name="Стол компьютерный", unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "tmc-requests"]),
            {
                "request_number": "17/TMC",
                "request_date": "2026-06-27",
                "status": "in_work",
                "comment": "Manual product",
                "item_product": [""],
                "item_name": ["Компьютерный стол"],
                "item_quantity": ["1"],
                "item_unit": ["шт."],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(TmcProduct.objects.filter(name="Компьютерный стол").exists())
        self.assertEqual(TmcProduct.objects.count(), 2)

    def test_request_create_form_uses_in_work_status_by_default(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("record_create", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")

        self.assertContains(response, 'hx-target="#table-area"')
        self.assertContains(response, "novalidate")
        self.assertContains(response, 'name="status"')
        self.assertContains(response, 'value="in_work"')
        self.assertNotContains(response, 'value="new"')
        self.assertNotIn(("new", "Новая"), TmcRequest._meta.get_field("status").choices)
        self.assertNotIn(("new", "Новая"), ACTIVE_NEED_STATUS_CHOICES)

    def test_request_number_field_has_autofocus_only_on_create_form(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            request_number="18-AF/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        self.client.login(username="operator", password="pass12345")

        create_response = self.client.get(reverse("record_create", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")
        update_response = self.client.get(reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(create_response, 'name="request_number"')
        self.assertContains(create_response, "autofocus")
        self.assertNotContains(update_response, "autofocus")

    def test_tmc_product_suggest_finds_words_in_any_order(self):
        TmcProduct.objects.create(name="Стол компьютерный", unit="шт.")
        TmcProduct.objects.create(name="Стол письменный", unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("tmc_product_suggest"), {"q": "компьютерный стол"})

        self.assertEqual(response.status_code, 200)
        names = [item["name"] for item in response.json()["results"]]
        self.assertEqual(names[0], "Стол компьютерный")

    def test_tmc_product_suggest_finds_typo_matches(self):
        TmcProduct.objects.create(name="Пылесос", unit="шт.")
        TmcProduct.objects.create(name="Пылесборник", unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("tmc_product_suggest"), {"q": "пылксос"})

        self.assertEqual(response.status_code, 200)
        names = [item["name"] for item in response.json()["results"]]
        self.assertEqual(names[0], "Пылесос")

    def test_tmc_product_suggest_requires_login(self):
        response = self.client.get(reverse("tmc_product_suggest"), {"q": "стол"})

        self.assertEqual(response.status_code, 302)

    def test_operator_cannot_access_foreign_organ_direct_urls(self):
        other_organ = TerritorialOrgan.objects.create(name="Foreign territorial organ", order_number=2)
        self.user.profile.allowed_organs.set([self.organ])
        foreign_request = TmcRequest.objects.create(
            territorial_organ=other_organ,
            created_by=self.user,
            request_number="FOREIGN-1",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=foreign_request, name="Desk", quantity=1, unit="шт.")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        foreign_photo = TerritorialOrganPhoto.objects.create(
            territorial_organ=other_organ,
            image=SimpleUploadedFile("foreign-organ.png", buffer.getvalue(), content_type="image/png"),
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.login(username="operator", password="pass12345")

        endpoints = [
            reverse("table_data", args=[other_organ.pk, "tmc-requests"]),
            reverse("record_create", args=[other_organ.pk, "tmc-requests"]),
            reverse("request_photos", args=[other_organ.pk, "tmc-requests", foreign_request.pk]),
            reverse("request_photos_download", args=[other_organ.pk, "tmc-requests", foreign_request.pk]),
            reverse("request_photo_picker", args=[other_organ.pk]),
            reverse("export_table", args=[other_organ.pk, "tmc-requests", "csv"]),
            reverse("photos", args=[other_organ.pk]),
            reverse("photo_download", args=[other_organ.pk, foreign_photo.pk]),
        ]

        for url in endpoints:
            with self.subTest(url=url):
                response = self.client.get(url, HTTP_HX_REQUEST="true")
                self.assertEqual(response.status_code, 404)

    def test_observer_can_view_table_but_cannot_write_records(self):
        User = get_user_model()
        observer = User.objects.create_user("observer", password="pass12345")
        UserProfile.objects.create(user=observer, role=UserProfile.Role.OBSERVER)
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            request_number="OBS-1",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Paper", quantity=1, unit="шт.")
        self.client.login(username="observer", password="pass12345")

        table_response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))
        create_response = self.client.get(reverse("record_create", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")
        update_response = self.client.get(reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")
        delete_response = self.client.post(reverse("record_delete", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")

        self.assertEqual(table_response.status_code, 200)
        self.assertContains(table_response, "OBS-1")
        self.assertNotContains(table_response, reverse("record_create", args=[self.organ.pk, "tmc-requests"]))
        self.assertNotContains(table_response, reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]))
        self.assertEqual(create_response.status_code, 404)
        self.assertEqual(update_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)

    def test_tmc_request_can_attach_and_show_photos(self):
        photo = self.create_photo("request-photo.png")
        photo.description = "Repair evidence"
        photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        form_response = self.client.get(reverse("record_create", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")
        self.assertContains(form_response, "Прикрепить фотографии")
        self.assertContains(form_response, "request-photo.png")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "tmc-requests"]),
            {
                "request_number": "15-Photo/TMC",
                "request_date": "2026-06-27",
                "status": "in_work",
                "comment": "With photo",
                "item_name": ["Desk"],
                "item_quantity": ["1"],
                "item_unit": ["шт."],
                "attached_photos": [str(photo.pk)],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        request_obj = TmcRequest.objects.get(request_number="15-Photo/TMC")
        self.assertTrue(RequestPhotoLink.objects.filter(photo=photo, object_id=request_obj.pk).exists())
        self.assertContains(response, "request-photo-count")
        self.assertContains(response, "Прикрепленные фотографии (1 шт.)")
        self.assertNotContains(response, "<span>1</span>", html=True)

        photos_response = self.client.get(reverse("request_photos", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(photos_response, "Repair evidence")
        self.assertContains(photos_response, "request-photo.png")
        self.assertContains(photos_response, "Открепить")
        self.assertContains(photos_response, "Скачать все")
        self.assertContains(photos_response, "Прикрепить еще")

        download_response = self.client.get(reverse("request_photos_download", args=[self.organ.pk, "tmc-requests", request_obj.pk]))
        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(download_response["Content-Type"], "application/zip")
        archive_data = b"".join(download_response.streaming_content)
        with zipfile.ZipFile(BytesIO(archive_data)) as archive:
            self.assertTrue(any(name.endswith(".png") for name in archive.namelist()))

    def test_request_photos_modal_can_replace_attached_photos(self):
        first = self.create_photo("first-proof.png")
        first.description = "First proof"
        first.save(update_fields=["description"])
        second = self.create_photo("second-proof.png")
        second.description = "Second proof"
        second.save(update_fields=["description"])
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="15-Replace/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Desk", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")
        self.client.post(
            reverse("request_photos", args=[self.organ.pk, "tmc-requests", request_obj.pk]),
            {"attached_photos": [str(first.pk)]},
            HTTP_HX_REQUEST="true",
        )
        self.assertTrue(RequestPhotoLink.objects.filter(photo=first, object_id=request_obj.pk).exists())

        response = self.client.post(
            reverse("request_photos", args=[self.organ.pk, "tmc-requests", request_obj.pk]),
            {"attached_photos": [str(second.pk)]},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(RequestPhotoLink.objects.filter(photo=first, object_id=request_obj.pk).exists())
        self.assertTrue(RequestPhotoLink.objects.filter(photo=second, object_id=request_obj.pk).exists())
        self.assertContains(response, "Second proof")
        self.assertContains(response, f'data-request-linked-photo="{second.pk}"')
        self.assertNotContains(response, f'data-request-linked-photo="{first.pk}"')
        self.assertIn("requestPhotosChanged", response["HX-Trigger"])

    def test_request_photos_ignore_photos_from_other_organ(self):
        other_organ = TerritorialOrgan.objects.create(name="Other photo organ", order_number=2)
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="15-CrossPhoto/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Desk", quantity=1, unit="шт.")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        foreign_photo = TerritorialOrganPhoto.objects.create(
            territorial_organ=other_organ,
            image=SimpleUploadedFile("foreign-request-photo.png", buffer.getvalue(), content_type="image/png"),
            description="Foreign evidence",
            created_by=self.user,
            updated_by=self.user,
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("request_photos", args=[self.organ.pk, "tmc-requests", request_obj.pk]),
            {"attached_photos": [str(foreign_photo.pk)]},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(RequestPhotoLink.objects.filter(photo=foreign_photo, object_id=request_obj.pk).exists())
        self.assertNotContains(response, "Foreign evidence")
        self.assertContains(response, "К заявке фотографии не прикреплены")

    def test_request_lightbox_groups_are_isolated_per_request(self):
        first = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            request_number="15-Lightbox-1/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        second = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            request_number="15-Lightbox-2/TMC",
            request_date="2026-06-28",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=first, name="Desk", quantity=1, unit="шт.")
        TmcRequestItem.objects.create(request=second, name="Chair", quantity=1, unit="шт.")
        first_photo = self.create_photo("first-lightbox.png")
        second_photo = self.create_photo("second-lightbox.png")
        content_type = ContentType.objects.get_for_model(TmcRequest, for_concrete_model=False)
        RequestPhotoLink.objects.create(
            territorial_organ=self.organ,
            photo=first_photo,
            content_type=content_type,
            object_id=first.pk,
            created_by=self.user,
        )
        RequestPhotoLink.objects.create(
            territorial_organ=self.organ,
            photo=second_photo,
            content_type=content_type,
            object_id=second.pk,
            created_by=self.user,
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))
        content = response.content.decode()
        first_group = f'request-{self.organ.pk}-tmc-requests-{first.pk}'
        second_group = f'request-{self.organ.pk}-tmc-requests-{second.pk}'

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'data-lightbox-group="{first_group}"')
        self.assertContains(response, f'data-lightbox-group="{second_group}"')
        self.assertContains(response, "first-lightbox.png")
        self.assertContains(response, "second-lightbox.png")
        self.assertGreaterEqual(content.count(f'data-lightbox-group="{first_group}"'), 2)
        self.assertGreaterEqual(content.count(f'data-lightbox-group="{second_group}"'), 2)

    def test_request_photo_picker_filters_paginates_and_keeps_selected(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Evidence")
        selected = self.create_photo("selected-proof.png")
        selected.description = "Already selected"
        selected.save(update_fields=["description"])
        folder_photo = self.create_photo("folder-proof.png")
        folder_photo.folder = folder
        folder_photo.description = "Folder proof"
        folder_photo.save(update_fields=["folder", "description"])
        root_photo = self.create_photo("root-proof.png")
        root_photo.description = "Root proof"
        root_photo.save(update_fields=["description"])
        for index in range(14):
            photo = self.create_photo(f"page-photo-{index}.png")
            photo.description = f"Page photo {index}"
            photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("request_photo_picker", args=[self.organ.pk]), {"attached_photos": [selected.pk], "photo_q": "folder"})
        self.assertContains(response, "selected-proof.png")
        self.assertContains(response, "folder-proof.png")
        self.assertNotContains(response, "root-proof.png")

        response = self.client.get(reverse("request_photo_picker", args=[self.organ.pk]), {"photo_folder": folder.pk})
        self.assertContains(response, "folder-proof.png")
        self.assertNotContains(response, "root-proof.png")

        response = self.client.get(reverse("request_photo_picker", args=[self.organ.pk]), {"photo_page": 2})
        self.assertContains(response, "request-photo-grid")
        self.assertContains(response, "photo_page=1")
        self.assertContains(response, 'class="pagination-jump"')
        self.assertContains(response, 'name="photo_page"')

    def test_tmc_status_history_records_status_changes(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="18/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Chair", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]),
            {
                "request_number": "18/TMC",
                "request_date": "2026-06-27",
                "status": "done",
                "due_date": "2026-06-29",
                "comment": "",
                "item_name": ["Chair"],
                "item_quantity": ["1"],
                "item_unit": ["pcs"],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.status_history(request_obj).filter(old_status="in_work", new_status="done", changed_by=self.user).exists())

    def test_tmc_edit_form_keeps_request_dates(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="18-1/TMC",
            request_date="2026-06-27",
            due_date="2026-06-28",
            status="done",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Chair", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(response, 'name="request_date"')
        self.assertContains(response, 'value="2026-06-27"')
        self.assertContains(response, 'name="due_date"')
        self.assertContains(response, 'value="2026-06-28"')

    def test_tmc_done_status_history_stores_completed_date(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="18-2/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Chair", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]),
            {
                "request_number": "18-2/TMC",
                "request_date": "2026-06-27",
                "status": "done",
                "due_date": "2026-06-29",
                "comment": "",
                "item_name": ["Chair"],
                "item_quantity": ["1"],
                "item_unit": ["pcs"],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        history = self.status_history(request_obj).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("tmc_status_history", args=[self.organ.pk, request_obj.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "29.06.2026")

    def test_tmc_status_history_modal_is_available(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="19/TMC",
            request_date="2026-06-27",
            status="done",
        )
        RequestStatusHistory.objects.create(
            content_type=ContentType.objects.get_for_model(request_obj, for_concrete_model=False),
            object_id=request_obj.pk,
            old_status="in_work",
            new_status="done",
            changed_by=self.user,
            note="Finished",
        )
        self.client.login(username="operator", password="pass12345")

        table_response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))
        self.assertContains(table_response, reverse("tmc_status_history", args=[self.organ.pk, request_obj.pk]))
        self.assertContains(table_response, "bi-clock-history")

        response = self.client.get(reverse("tmc_status_history", args=[self.organ.pk, request_obj.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(response, "История изменений статуса заявки 19/TMC")
        self.assertContains(response, request_obj.get_status_display())
        self.assertContains(response, "Finished")

    def test_status_history_button_is_hidden_without_history(self):
        request_without_history = TmcRequest.objects.create(
            territorial_organ=self.organ,
            request_number="18/TMC",
            request_date="2026-06-26",
            status="in_work",
        )
        request_with_history = TmcRequest.objects.create(
            territorial_organ=self.organ,
            request_number="19/TMC",
            request_date="2026-06-27",
            status="done",
        )
        RequestStatusHistory.objects.create(
            content_type=ContentType.objects.get_for_model(request_with_history, for_concrete_model=False),
            object_id=request_with_history.pk,
            old_status="in_work",
            new_status="done",
            changed_by=self.user,
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertNotContains(response, reverse("tmc_status_history", args=[self.organ.pk, request_without_history.pk]))
        self.assertContains(response, reverse("tmc_status_history", args=[self.organ.pk, request_with_history.pk]))

    def test_tmc_xlsx_export_has_grouped_document_layout(self):
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="20/TMC", request_date="2026-06-27", status="in_work", comment="First comment")
        TmcRequestItem.objects.create(request=first, name="Desk", quantity=5, unit="pcs")
        TmcRequestItem.objects.create(request=first, name="Chair", quantity=5, unit="pcs")
        second = TmcRequest.objects.create(territorial_organ=self.organ, request_number="21/TMC", request_date="2026-06-26", status="in_work", comment="Second comment")
        TmcRequestItem.objects.create(request=second, name="Keyboard", quantity=3, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("export_table", args=[self.organ.pk, "tmc-requests", "xlsx"]))

        self.assertEqual(response.status_code, 200)
        workbook = self.response_workbook(response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Сведения о потребности ТМЦ")
        self.assertEqual(sheet["C1"].value, "Заявка")
        self.assertEqual(sheet["F1"].value, "Описание")
        self.assertIn("A1:B1", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("C1:E1", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("F1:F2", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("C3:C4", {str(item) for item in sheet.merged_cells.ranges})
        self.assertEqual(sheet["A3"].value, "Desk")
        self.assertEqual(sheet["A4"].value, "Chair")
        self.assertEqual(sheet["C3"].value, "20/TMC")
        self.assertEqual(sheet["E3"].value, first.get_status_display())
        self.assertEqual(sheet["F3"].value, "First comment")
        self.assertEqual(sheet["B3"].border.right.style, "medium")
        self.assertEqual(sheet["E3"].border.right.style, "medium")
        self.assertEqual(sheet["F3"].border.right.style, "medium")
        self.assertEqual(sheet["A4"].border.bottom.style, "medium")

    def test_tmc_filters_by_status_date_range_and_text(self):
        matching = TmcRequest.objects.create(territorial_organ=self.organ, request_number="22/TMC", request_date="2026-06-20", status="in_work", comment="Office")
        TmcRequestItem.objects.create(request=matching, name="Monitor", quantity=2, unit="pcs")
        wrong_status = TmcRequest.objects.create(territorial_organ=self.organ, request_number="23/TMC", request_date="2026-06-20", status="done", comment="Office")
        TmcRequestItem.objects.create(request=wrong_status, name="Monitor", quantity=1, unit="pcs")
        wrong_date = TmcRequest.objects.create(territorial_organ=self.organ, request_number="24/TMC", request_date="2026-05-20", status="in_work", comment="Office")
        TmcRequestItem.objects.create(request=wrong_date, name="Monitor", quantity=1, unit="pcs")
        wrong_text = TmcRequest.objects.create(territorial_organ=self.organ, request_number="25/TMC", request_date="2026-06-20", status="in_work", comment="Warehouse")
        TmcRequestItem.objects.create(request=wrong_text, name="Printer", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "Monitor"},
        )

        self.assertContains(response, "22/TMC")
        self.assertNotContains(response, "23/TMC")
        self.assertNotContains(response, "24/TMC")
        self.assertNotContains(response, "25/TMC")
        self.assertContains(response, "status=in_work")
        self.assertContains(response, "date_from=2026-06-01")
        self.assertContains(response, "date_to=2026-06-30")
        self.assertContains(response, "q=Monitor")
        self.assertContains(response, "В работе")
        self.assertNotContains(response, 'value="new"')
        self.assertNotContains(response, "Новых")
        self.assertContains(response, "Исполнено")
        self.assertContains(response, "Отклонено")
        self.assertContains(response, "<strong>1</strong>", html=True)

    def test_zero_status_summary_pills_are_muted_consistently(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, "summary-pill-in-work is-zero")
        self.assertContains(response, "summary-pill-done is-zero")
        self.assertContains(response, "summary-pill-rejected is-zero")

    def test_tmc_table_supports_multi_organ_summary_mode(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="40/TMC", request_date="2026-06-20", status="in_work", comment="Office")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="41/TMC", request_date="2026-06-21", status="in_work", comment="Office")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Бумага А4", quantity=7, unit="пач.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "q": "Бумага А4"},
        )

        self.assertContains(response, "Территориальный орган")
        self.assertContains(response, "Test territorial organ")
        self.assertContains(response, "Other territorial organ")
        self.assertContains(response, "40/TMC")
        self.assertContains(response, "41/TMC")
        self.assertContains(response, f'name="organ_ids" value="{self.organ.pk}"')
        self.assertContains(response, f'name="organ_ids" value="{other_organ.pk}"')
        self.assertNotContains(response, "Добавить")

    def test_department_panel_preserves_multi_organ_querystring(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("department_tables", args=[self.organ.pk, "tmc"]),
            {"organ_ids": [self.organ.pk, other_organ.pk]},
        )

        self.assertContains(response, "Сводный просмотр: 2 территориальных органов")
        self.assertContains(response, f"organ_ids={self.organ.pk}")
        self.assertContains(response, f"organ_ids={other_organ.pk}")

    def test_department_panel_restores_requested_table_and_filters(self):
        Department.objects.create(name="Transport", slug="transport", order_number=2)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("department_tables", args=[self.organ.pk, "transport"]),
            {"table": "vehicle-fuel", "status": "in_work", "q": "diesel"},
        )

        self.assertContains(response, 'data-table-key="vehicle-fuel"')
        self.assertContains(response, reverse("table_data", args=[self.organ.pk, "vehicle-fuel"]) + "?status=in_work&amp;q=diesel")
        self.assertContains(response, 'data-table-key="vehicle-fuel" hx-get')
        self.assertNotContains(response, reverse("table_data", args=[self.organ.pk, "vehicle-repair"]) + "?status=in_work")

    def test_multi_organ_summary_keeps_row_actions_for_writable_organs(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="42/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="43/TMC", request_date="2026-06-21", status="in_work")
        TmcRequestItem.objects.create(request=first, name="Paper", quantity=5, unit="pcs")
        TmcRequestItem.objects.create(request=second, name="Paper", quantity=7, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "q": "Paper"},
        )

        self.assertContains(response, reverse("record_update", args=[self.organ.pk, "tmc-requests", first.pk]))
        self.assertContains(response, reverse("record_update", args=[other_organ.pk, "tmc-requests", second.pk]))
        self.assertContains(response, reverse("record_delete", args=[self.organ.pk, "tmc-requests", first.pk]))
        self.assertContains(response, reverse("record_delete", args=[other_organ.pk, "tmc-requests", second.pk]))
        self.assertNotContains(response, reverse("record_create", args=[self.organ.pk, "tmc-requests"]))

    def test_operator_can_write_only_assigned_departments(self):
        transport = Department.objects.create(name="Transport", slug="transport", order_number=2)
        self.user.profile.allowed_departments.add(transport)
        request_obj = TmcRequest.objects.create(territorial_organ=self.organ, request_number="43/TMC", request_date="2026-06-20", status="in_work")
        TmcRequestItem.objects.create(request=request_obj, name="Paper", quantity=5, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        table_response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))
        create_response = self.client.get(reverse("record_create", args=[self.organ.pk, "tmc-requests"]), HTTP_HX_REQUEST="true")
        update_response = self.client.get(reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(table_response, "43/TMC")
        self.assertNotContains(table_response, reverse("record_create", args=[self.organ.pk, "tmc-requests"]))
        self.assertNotContains(table_response, reverse("record_update", args=[self.organ.pk, "tmc-requests", request_obj.pk]))
        self.assertEqual(create_response.status_code, 404)
        self.assertEqual(update_response.status_code, 404)

    def test_tmc_table_can_group_products_across_selected_organs(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="44/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="45/TMC", request_date="2026-06-21", status="in_work")
        third = TmcRequest.objects.create(territorial_organ=other_organ, request_number="46/TMC", request_date="2026-06-22", status="done")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Бумага А4", quantity=7, unit="пач.")
        TmcRequestItem.objects.create(request=third, name="Кресло офисное", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "products"},
        )

        self.assertContains(response, "По заявкам")
        self.assertContains(response, "По ТМЦ")
        self.assertContains(response, "По территориальному органу")
        self.assertContains(response, "По дате")
        self.assertContains(response, 'name="group"')
        self.assertContains(response, "Бумага А4")
        self.assertContains(response, "Кресло офисное")
        self.assertContains(response, "tmc-drilldown-link")
        self.assertContains(response, f"organ_ids={self.organ.pk}")
        self.assertContains(response, f"organ_ids={other_organ.pk}")
        self.assertContains(response, f"q={quote_plus('Бумага А4')}")
        self.assertContains(response, "<td class=\"text-center\">2</td>", html=True)
        self.assertContains(response, "<td class=\"text-center\">12</td>", html=True)
        self.assertContains(response, "позиций")
        self.assertContains(response, "Применены фильтры:")
        self.assertContains(response, "выборочно: 2 органов")
        self.assertContains(response, "группировка: По ТМЦ")
        self.assertContains(response, "Сбросить все")
        self.assertContains(response, "data-reset-table-state")
        self.assertContains(response, "Позиций найдено")
        self.assertContains(response, "Всего заявок")
        self.assertContains(response, "Всего органов")
        self.assertContains(response, "Общее количество")
        self.assertContains(response, 'data-download-preparing="Подготовка экспорта..."')
        self.assertContains(response, "<strong>2</strong>", count=3, html=True)
        self.assertContains(response, "<strong>3</strong>", html=True)
        self.assertContains(response, "<strong>13</strong>", html=True)
        self.assertNotContains(response, "summary-pill-in-work")
        self.assertNotContains(response, "summary-pill-new")
        self.assertNotContains(response, "summary-pill-done")
        self.assertNotContains(response, "summary-pill-rejected")
        self.assertNotContains(response, "Сбросить фильтры")
        self.assertNotContains(response, reverse("record_update", args=[self.organ.pk, "tmc-requests", first.pk]))

    def test_tmc_table_can_group_by_territorial_organs_when_multiple_selected(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="51/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="52/TMC", request_date="2026-06-21", status="in_work")
        third = TmcRequest.objects.create(territorial_organ=other_organ, request_number="53/TMC", request_date="2026-06-22", status="done")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Бумага А4", quantity=7, unit="пач.")
        TmcRequestItem.objects.create(request=third, name="Кресло офисное", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "organs"},
        )

        self.assertContains(response, "Территориальный орган")
        self.assertContains(response, "Test territorial organ")
        self.assertContains(response, "Other territorial organ")
        self.assertContains(response, "Заявок")
        self.assertContains(response, "Позиций ТМЦ")
        self.assertContains(response, "Общее количество")
        self.assertContains(response, "группировка: По территориальному органу")
        self.assertContains(response, "органов")
        self.assertContains(response, "Органов найдено")
        self.assertContains(response, "<td class=\"text-center\">2</td>", html=True)
        self.assertNotContains(response, "tmc-drilldown-link")
        self.assertContains(response, "summary-pill-in-work")

    def test_tmc_organ_grouping_is_available_only_for_multi_organ_mode(self):
        request_obj = TmcRequest.objects.create(territorial_organ=self.organ, request_number="54/TMC", request_date="2026-06-20", status="in_work")
        TmcRequestItem.objects.create(request=request_obj, name="Сканер", quantity=2, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]), {"group": "organs"})

        self.assertContains(response, "54/TMC")
        self.assertContains(response, "По территориальному органу")
        self.assertContains(response, "disabled")
        self.assertNotContains(response, "группировка: По территориальному органу")

    def test_tmc_table_can_group_by_request_date(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="57/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="58/TMC", request_date="2026-06-20", status="in_work")
        third = TmcRequest.objects.create(territorial_organ=other_organ, request_number="59/TMC", request_date="2026-06-21", status="done")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Кресло офисное", quantity=2, unit="шт.")
        TmcRequestItem.objects.create(request=third, name="Сканер", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "dates"},
        )

        self.assertContains(response, "Дата")
        self.assertContains(response, "20.06.2026")
        self.assertContains(response, "21.06.2026")
        self.assertContains(response, "Территориальных органов")
        self.assertContains(response, "группировка: По дате")
        self.assertContains(response, "дней")
        self.assertContains(response, "Дней найдено")
        self.assertContains(response, "<td class=\"text-center\">2</td>", html=True)
        self.assertNotContains(response, "tmc-drilldown-link")
        self.assertContains(response, "summary-pill-in-work")

    def test_table_active_conditions_show_filters_and_reset(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            request_number="47/TMC",
            request_date="2026-06-20",
            status="in_work",
            comment="Office paper",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Бумага А4", quantity=5, unit="пач.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"q": "бумага", "status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30"},
        )

        self.assertContains(response, "Применены фильтры:")
        self.assertContains(response, "поиск: бумага")
        self.assertContains(response, "исполнение: В работе")
        self.assertContains(response, "с 01.06.2026")
        self.assertContains(response, "по 30.06.2026")
        self.assertContains(response, "Сбросить все")
        self.assertContains(response, "data-reset-table-state")
        self.assertContains(response, 'hx-target="#workspace"')
        self.assertContains(response, f'{reverse("department_tables", args=[self.organ.pk, "tmc"])}?table=tmc-requests')
        self.assertNotContains(response, "Сбросить фильтры")

    def test_request_table_search_triggers_while_typing(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, 'id="table-search-tmc-requests"')
        self.assertContains(response, "input delay:500ms from:#table-search-tmc-requests")
        self.assertContains(response, "change")
        self.assertNotContains(response, "from:input")

    def test_tmc_search_is_case_insensitive_for_cyrillic(self):
        matching = TmcRequest.objects.create(territorial_organ=self.organ, request_number="32/TMC", request_date="2026-06-20", status="in_work", comment="Склад")
        TmcRequestItem.objects.create(request=matching, name="Стол письменный", quantity=2, unit="шт.")
        other = TmcRequest.objects.create(territorial_organ=self.organ, request_number="33/TMC", request_date="2026-06-20", status="in_work", comment="Кабинет")
        TmcRequestItem.objects.create(request=other, name="Кресло офисное", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]), {"q": "стол"})

        self.assertContains(response, "32/TMC")
        self.assertContains(response, "Стол письменный")
        self.assertNotContains(response, "33/TMC")

    def test_tmc_in_work_counter_ignores_selected_status_filter(self):
        in_work = TmcRequest.objects.create(territorial_organ=self.organ, request_number="28/TMC", request_date="2026-06-20", status="in_work", comment="Office")
        TmcRequestItem.objects.create(request=in_work, name="Monitor", quantity=1, unit="pcs")
        done = TmcRequest.objects.create(territorial_organ=self.organ, request_number="29/TMC", request_date="2026-06-20", status="done", comment="Office")
        TmcRequestItem.objects.create(request=done, name="Monitor", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "tmc-requests"]),
            {"status": "done", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "Monitor"},
        )

        self.assertContains(response, "29/TMC")
        self.assertNotContains(response, "28/TMC")
        self.assertContains(response, "В работе")
        self.assertContains(response, "Исполнено")
        self.assertContains(response, "<strong>1</strong>", html=True)

    def test_tmc_date_filters_have_default_range(self):
        today = timezone.localdate()
        oldest_date = today - timedelta(days=10)
        oldest = TmcRequest.objects.create(territorial_organ=self.organ, request_number="30/TMC", request_date=oldest_date, status="in_work")
        TmcRequestItem.objects.create(request=oldest, name="Archive box", quantity=1, unit="pcs")
        future = TmcRequest.objects.create(territorial_organ=self.organ, request_number="31/TMC", request_date=today + timedelta(days=1), status="in_work")
        TmcRequestItem.objects.create(request=future, name="Future item", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, f'value="{oldest_date.isoformat()}"')
        self.assertContains(response, f'value="{today.isoformat()}"')
        self.assertContains(response, f'data-default-date-from="{oldest_date.isoformat()}"')
        self.assertContains(response, f'data-default-date-to="{today.isoformat()}"')
        self.assertContains(response, "30/TMC")
        self.assertNotContains(response, "31/TMC")

    def test_tmc_date_filters_default_to_today_without_records(self):
        today = timezone.localdate()
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, f'name="date_from" value="{today.isoformat()}"')
        self.assertContains(response, f'name="date_to" value="{today.isoformat()}"')
        self.assertContains(response, f'data-default-date-from="{today.isoformat()}"')
        self.assertContains(response, f'data-default-date-to="{today.isoformat()}"')

    def test_table_pagination_uses_photo_style_controls_above_table(self):
        for index in range(21):
            request_obj = TmcRequest.objects.create(
                territorial_organ=self.organ,
                request_number=f"PAGE-{index:02d}",
                request_date="2026-06-20",
                status="in_work",
            )
            TmcRequestItem.objects.create(request=request_obj, name="Paper", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, 'class="table-pagination"')
        self.assertContains(response, 'class="photo-page-number is-active"')
        self.assertContains(response, "page=2")
        self.assertContains(response, 'class="pagination-jump"')
        self.assertContains(response, 'name="page"')
        self.assertContains(response, f"/ {response.context['page'].paginator.num_pages}")
        self.assertNotContains(response, "pagination-jump-submit")
        self.assertNotContains(response, "btn-group btn-group-sm")

    def test_tmc_xlsx_export_uses_current_filters(self):
        included = TmcRequest.objects.create(territorial_organ=self.organ, request_number="26/TMC", request_date="2026-06-20", status="in_work")
        TmcRequestItem.objects.create(request=included, name="Scanner", quantity=1, unit="pcs")
        excluded = TmcRequest.objects.create(territorial_organ=self.organ, request_number="27/TMC", request_date="2026-06-20", status="done")
        TmcRequestItem.objects.create(request=excluded, name="Scanner", quantity=1, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("export_table", args=[self.organ.pk, "tmc-requests", "xlsx"]), {"status": "in_work", "q": "Scanner", "download_token": "exporttest"})

        self.assertIn("download-ready-exporttest", response.cookies)
        workbook = self.response_workbook(response)
        sheet = workbook.active
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("26/TMC", values)
        self.assertNotIn("27/TMC", values)

    def test_tmc_xlsx_export_includes_organ_column_for_multi_organ_mode(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="28/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="29/TMC", request_date="2026-06-21", status="done")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=first, name="Папка-регистратор", quantity=2, unit="шт.")
        TmcRequestItem.objects.create(request=second, name="Кресло офисное", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "tmc-requests", "xlsx"]),
            {"organ_ids": [self.organ.pk, other_organ.pk]},
        )

        workbook = self.response_workbook(response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Территориальный орган")
        self.assertEqual(sheet["B1"].value, "Сведения о потребности ТМЦ")
        self.assertEqual(sheet["D1"].value, "Заявка")
        self.assertEqual(sheet["G1"].value, "Описание")
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("Test territorial organ", values)
        self.assertIn("Other territorial organ", values)
        self.assertIn("28/TMC", values)
        self.assertIn("29/TMC", values)
        self.assertIn("Бумага А4", values)
        self.assertIn("Папка-регистратор", values)
        self.assertIn("Кресло офисное", values)

    def test_tmc_grouped_xlsx_export_matches_grouped_table(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="48/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="49/TMC", request_date="2026-06-21", status="in_work")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Бумага А4", quantity=7, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Кресло офисное", quantity=1, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "tmc-requests", "xlsx"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "products"},
        )

        workbook = self.response_workbook(response)
        sheet = workbook.active
        self.assertEqual(sheet.title, "ТМЦ")
        self.assertEqual([sheet.cell(row=1, column=column).value for column in range(1, 6)], ["Наименование ТМЦ", "Заявок", "Территориальных органов", "Общее количество", "Единица измерения"])
        values = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertIn("Бумага А4", values)
        self.assertIn("Кресло офисное", values)
        self.assertIn(12, values)
        self.assertNotIn("48/TMC", values)
        self.assertNotIn("49/TMC", values)

    def test_tmc_grouped_csv_export_matches_grouped_table(self):
        request_obj = TmcRequest.objects.create(territorial_organ=self.organ, request_number="50/TMC", request_date="2026-06-20", status="in_work")
        TmcRequestItem.objects.create(request=request_obj, name="Сканер", quantity=2, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("export_table", args=[self.organ.pk, "tmc-requests", "csv"]), {"group": "products"})

        rows = list(csv.reader(self.response_bytes(response).decode("utf-8-sig").splitlines()))
        self.assertEqual(rows[0], ["Наименование ТМЦ", "Заявок", "Общее количество", "Единица измерения"])
        self.assertEqual(rows[1], ["Сканер", "1", "2", "шт."])
        self.assertNotIn("50/TMC", ",".join(rows[1]))

    def test_tmc_organ_grouped_csv_export_matches_grouped_table(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="55/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="56/TMC", request_date="2026-06-21", status="in_work")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Кресло офисное", quantity=2, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "tmc-requests", "csv"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "organs"},
        )

        rows = list(csv.reader(self.response_bytes(response).decode("utf-8-sig").splitlines()))
        self.assertEqual(rows[0], ["Территориальный орган", "Заявок", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"])
        self.assertIn(["Test territorial organ", "1", "1", "5", "1", "0", "0"], rows)
        self.assertIn(["Other territorial organ", "1", "1", "2", "1", "0", "0"], rows)
        self.assertNotIn("55/TMC", ",".join(",".join(row) for row in rows))

    def test_tmc_date_grouped_csv_export_matches_grouped_table(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        first = TmcRequest.objects.create(territorial_organ=self.organ, request_number="60/TMC", request_date="2026-06-20", status="in_work")
        second = TmcRequest.objects.create(territorial_organ=other_organ, request_number="61/TMC", request_date="2026-06-20", status="in_work")
        TmcRequestItem.objects.create(request=first, name="Бумага А4", quantity=5, unit="пач.")
        TmcRequestItem.objects.create(request=second, name="Кресло офисное", quantity=2, unit="шт.")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "tmc-requests", "csv"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "dates"},
        )

        rows = list(csv.reader(self.response_bytes(response).decode("utf-8-sig").splitlines()))
        self.assertEqual(rows[0], ["Дата", "Заявок", "Территориальных органов", "Позиций ТМЦ", "Общее количество", "В работе", "Исполнено", "Отклонено"])
        self.assertIn(["20.06.2026", "2", "2", "2", "7", "2", "0", "0"], rows)
        self.assertNotIn("60/TMC", ",".join(",".join(row) for row in rows))

    def test_citsizi_filter_by_equipment_type(self):
        CitsiziEquipment.objects.create(territorial_organ=self.organ, request_number="C-1", request_date="2026-06-20", equipment_type="communication", quantity=1)
        CitsiziEquipment.objects.create(territorial_organ=self.organ, request_number="C-2", request_date="2026-06-20", equipment_type="computing", quantity=1)
        self.client.login(username="operator", password="pass12345")
        response = self.client.get(reverse("table_data", args=[self.organ.pk, "citsizi-equipment"]), {"equipment_type": "communication"})
        self.assertContains(response, "C-1")
        self.assertContains(response, "Средства связи")
        self.assertNotContains(response, "C-2")

    def test_citsizi_form_includes_sound_alert_equipment_type(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("record_create", args=[self.organ.pk, "citsizi-equipment"]), HTTP_HX_REQUEST="true")

        self.assertContains(response, 'hx-target="#table-area"')
        self.assertContains(response, "novalidate")
        self.assertEqual(response.content.decode().count("<form"), 1)
        self.assertNotContains(response, '<form class="pagination-jump"', html=False)
        self.assertContains(response, "Выберите тип техники")
        self.assertNotContains(response, "---------")
        self.assertContains(response, f'value="{EquipmentType.SOUND_ALERT}"')

    def test_citsizi_equipment_type_is_required(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "citsizi-equipment"]),
            {
                "request_number": "C-empty",
                "request_date": "2026-06-20",
                "quantity": "1",
                "status": "in_work",
                "equipment_type": "",
                "comment": "",
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Выберите тип техники")
        self.assertContains(response, 'hx-target="#table-area"')
        self.assertEqual(response["HX-Retarget"], "#modal-content")
        self.assertIn("equipment_type", response.context["form"].errors)
        self.assertFalse(CitsiziEquipment.objects.filter(request_number="C-empty").exists())

    def test_citsizi_valid_create_retargets_table_area(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "citsizi-equipment"]),
            {
                "request_number": "C-valid",
                "request_date": "2026-06-20",
                "quantity": "1",
                "status": "in_work",
                "equipment_type": EquipmentType.COMMUNICATION,
                "comment": "",
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("HX-Retarget", response)
        self.assertTrue(CitsiziEquipment.objects.filter(request_number="C-valid").exists())
        self.assertContains(response, "C-valid")

    def test_citsizi_request_table_history_filters_and_styled_export(self):
        included = CitsiziEquipment.objects.create(territorial_organ=self.organ, request_number="C-10", request_date="2026-06-20", equipment_type="communication", quantity=3, status="in_work", comment="Install radio")
        excluded = CitsiziEquipment.objects.create(territorial_organ=self.organ, request_number="C-11", request_date="2026-06-20", equipment_type="computing", quantity=2, status="done")
        self.create_status_history_entry(included)
        self.client.login(username="operator", password="pass12345")

        table_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "citsizi-equipment"]),
            {"status": "in_work", "equipment_type": "communication", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "C-10"},
        )
        self.assertContains(table_response, "<th>Номер</th>", html=True)
        self.assertContains(table_response, "<th>Дата</th>", html=True)
        self.assertContains(table_response, "<th>Количество</th>", html=True)
        self.assertContains(table_response, "<th>Исполнение</th>", html=True)
        self.assertContains(table_response, "<th>Тип техники</th>", html=True)
        self.assertContains(table_response, "<th>Описание</th>", html=True)
        self.assertContains(table_response, "Install radio")
        self.assertContains(table_response, included.request_number)
        self.assertNotContains(table_response, excluded.request_number)
        self.assertContains(table_response, "equipment_type=communication")
        self.assertContains(table_response, "bi-clock-history")

        update_response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "citsizi-equipment", included.pk]),
            {
                "request_number": "C-10",
                "request_date": "2026-06-20",
                "quantity": "3",
                "status": "done",
                "equipment_type": "communication",
                "due_date": "2026-06-29",
                "comment": "Install radio completed",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(update_response.status_code, 200)
        history = self.status_history(included).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("citsizi_status_history", args=[self.organ.pk, included.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки C-10")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "citsizi-equipment", "xlsx"]), {"status": "done", "equipment_type": "communication"})
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Номер")
        self.assertEqual(sheet["E1"].value, "Тип техники")
        self.assertEqual(sheet["F1"].value, "Описание")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["F1"].border.right.style, "medium")

    def test_regular_table_headers_start_with_capital_letter(self):
        VehicleInventory.objects.create(territorial_organ=self.organ, required_count=5, available_count=4, broken_count=1, writeoff_count=0)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "vehicle-inventory"]))

        self.assertContains(response, "<th>Положено</th>", html=True)
        self.assertNotContains(response, "<th>положено</th>", html=True)

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "vehicle-inventory", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["B1"].value, "Положено")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00D6EAF7")
        self.assertEqual(sheet["A1"].border.bottom.style, "medium")

    def test_vehicle_inventory_has_date_as_first_column(self):
        VehicleInventory.objects.create(
            territorial_organ=self.organ,
            state_date="2026-06-27",
            required_count=5,
            available_count=4,
            broken_count=1,
            writeoff_count=0,
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "vehicle-inventory"]))

        self.assertContains(response, "<th>Дата</th>", html=True)
        self.assertContains(response, "table-vehicle-inventory")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "vehicle-inventory", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Дата")
        self.assertEqual(sheet["B1"].value, "Положено")
        self.assertEqual(sheet.column_dimensions["E"].width, 38)
        self.assertEqual(sheet["A2"].alignment.horizontal, "center")
        self.assertEqual(sheet["E2"].border.right.style, "medium")

    def test_vehicle_repair_request_shows_comment_column(self):
        request_obj = VehicleRepairRequest.objects.create(
            territorial_organ=self.organ,
            request_number="R-1",
            request_date="2026-06-27",
            status="in_work",
            comment="Needs diagnostics",
        )
        self.create_status_history_entry(request_obj)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "vehicle-repair"]))

        self.assertNotContains(response, "<th>Дата исполнения заявки</th>", html=True)
        self.assertContains(response, "<th>Описание</th>", html=True)
        self.assertContains(response, "Needs diagnostics")
        self.assertContains(response, "table-vehicle-repair")
        self.assertContains(response, "table-row-actions")
        self.assertContains(response, "bi-clock-history")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "vehicle-repair", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["D1"].value, "Описание")
        self.assertEqual(sheet["D2"].value, "Needs diagnostics")
        self.assertIsNone(sheet["E1"].value)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["D1"].fill.fgColor.rgb, "00D6EAF7")
        self.assertEqual(sheet["D1"].border.right.style, "medium")
        self.assertEqual(sheet["A2"].alignment.horizontal, "center")
        self.assertIsNone(sheet["D2"].alignment.horizontal)

    def test_request_date_defaults_to_today_in_create_form(self):
        today = timezone.localdate()
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("record_create", args=[self.organ.pk, "vehicle-repair"]), HTTP_HX_REQUEST="true")

        self.assertContains(response, 'name="request_date"')
        self.assertContains(response, f'value="{today.isoformat()}"')
        self.assertContains(response, 'name="completed_at"')

    def test_vehicle_repair_status_history_records_completed_date(self):
        request_obj = VehicleRepairRequest.objects.create(
            territorial_organ=self.organ,
            request_number="R-2",
            request_date="2026-06-27",
            status="in_work",
            comment="Initial",
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "vehicle-repair", request_obj.pk]),
            {
                "request_number": "R-2",
                "request_date": "2026-06-27",
                "status": "done",
                "completed_at": "2026-06-29",
                "comment": "Completed",
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        history = self.status_history(request_obj).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("vehicle_repair_status_history", args=[self.organ.pk, request_obj.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки R-2")
        self.assertContains(modal, "Дата исполнения заявки")
        self.assertContains(modal, "29.06.2026")

    def test_vehicle_repair_filters_by_status_date_range_and_text(self):
        matching = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-10", request_date="2026-06-20", status="in_work", comment="Diagnostics")
        wrong_status = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-11", request_date="2026-06-20", status="done", comment="Diagnostics")
        wrong_date = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-12", request_date="2026-05-20", status="in_work", comment="Diagnostics")
        wrong_text = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-13", request_date="2026-06-20", status="in_work", comment="Oil")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "vehicle-repair"]),
            {"status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "Diagnostics"},
        )

        self.assertContains(response, matching.request_number)
        self.assertNotContains(response, wrong_status.request_number)
        self.assertNotContains(response, wrong_date.request_number)
        self.assertNotContains(response, wrong_text.request_number)
        self.assertContains(response, "status=in_work")
        self.assertContains(response, "date_from=2026-06-01")
        self.assertContains(response, "date_to=2026-06-30")
        self.assertContains(response, "q=Diagnostics")
        self.assertContains(response, "В работе")
        self.assertNotContains(response, 'value="new"')
        self.assertNotContains(response, "Новых")
        self.assertContains(response, "Исполнено")
        self.assertContains(response, "Отклонено")

    def test_vehicle_repair_xlsx_export_uses_current_filters(self):
        included = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-20", request_date="2026-06-20", status="in_work", comment="Transmission")
        excluded = VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-21", request_date="2026-06-20", status="done", comment="Transmission")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("export_table", args=[self.organ.pk, "vehicle-repair", "xlsx"]), {"status": "in_work", "q": "Transmission"})

        workbook = self.response_workbook(response)
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertIn(included.request_number, values)
        self.assertNotIn(excluded.request_number, values)

    def test_request_tables_support_date_and_organ_grouping(self):
        other_organ = TerritorialOrgan.objects.create(name="Other territorial organ", order_number=2)
        VehicleRepairRequest.objects.create(territorial_organ=self.organ, request_number="R-30", request_date="2026-06-20", status="in_work", comment="Diagnostics")
        VehicleRepairRequest.objects.create(territorial_organ=other_organ, request_number="R-31", request_date="2026-06-20", status="done", comment="Diagnostics")
        VehicleRepairRequest.objects.create(territorial_organ=other_organ, request_number="R-32", request_date="2026-06-21", status="rejected", comment="Oil")
        self.client.login(username="operator", password="pass12345")

        date_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "vehicle-repair"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "dates"},
        )
        self.assertContains(date_response, "По заявкам")
        self.assertContains(date_response, "По дате")
        self.assertContains(date_response, "По территориальному органу")
        self.assertNotContains(date_response, "По ТМЦ")
        self.assertContains(date_response, "20.06.2026")
        self.assertContains(date_response, "21.06.2026")
        self.assertContains(date_response, "В работе")
        self.assertContains(date_response, "Исполнено")
        self.assertContains(date_response, "Отклонено")
        self.assertContains(date_response, "группировка: По дате")
        self.assertNotContains(date_response, "R-30")

        organ_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "vehicle-repair"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "organs"},
        )
        self.assertContains(organ_response, "Test territorial organ")
        self.assertContains(organ_response, "Other territorial organ")
        self.assertContains(organ_response, "группировка: По территориальному органу")
        self.assertNotContains(organ_response, "R-31")

        export_response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "vehicle-repair", "csv"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "group": "dates"},
        )
        rows = list(csv.reader(self.response_bytes(export_response).decode("utf-8-sig").splitlines()))
        self.assertEqual(rows[0], ["Дата", "Заявок", "Территориальных органов", "В работе", "Исполнено", "Отклонено"])
        self.assertIn(["20.06.2026", "2", "2", "1", "1", "0"], rows)
        self.assertNotIn("R-30", ",".join(",".join(row) for row in rows))

    def test_vehicle_fuel_request_matches_vehicle_repair_table_behavior(self):
        request_obj = VehicleFuelRequest.objects.create(
            territorial_organ=self.organ,
            request_number="GSM-1",
            request_date="2026-06-27",
            status="in_work",
            comment="Fuel cards",
        )
        self.create_status_history_entry(request_obj)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "vehicle-fuel"]))

        self.assertContains(response, "<th>Описание</th>", html=True)
        self.assertContains(response, "Fuel cards")
        self.assertContains(response, "table-vehicle-fuel")
        self.assertContains(response, "bi-clock-history")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "vehicle-fuel", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Номер")
        self.assertEqual(sheet["D1"].value, "Описание")
        self.assertEqual(sheet["D2"].value, "Fuel cards")
        self.assertEqual(sheet["D1"].border.right.style, "medium")

    def test_vehicle_fuel_status_history_records_completed_date(self):
        request_obj = VehicleFuelRequest.objects.create(
            territorial_organ=self.organ,
            request_number="GSM-2",
            request_date="2026-06-27",
            status="in_work",
            comment="Initial",
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "vehicle-fuel", request_obj.pk]),
            {
                "request_number": "GSM-2",
                "request_date": "2026-06-27",
                "status": "done",
                "completed_at": "2026-06-29",
                "comment": "Completed",
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        history = self.status_history(request_obj).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("vehicle_fuel_status_history", args=[self.organ.pk, request_obj.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки GSM-2")
        self.assertContains(modal, "Дата исполнения заявки")
        self.assertContains(modal, "29.06.2026")

    def test_fire_inventory_tabs_have_date_short_headers_and_styled_export(self):
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date="2026-06-27", required_count=10, available_count=8, expiry_date="2026-12-31", writeoff_count=1)
        FireAlarm.objects.create(territorial_organ=self.organ, state_date="2026-06-27", required_objects=5, equipped_objects=4, broken_objects=1)
        SecurityAlarm.objects.create(territorial_organ=self.organ, state_date="2026-06-27", required_objects=6, equipped_objects=5, broken_objects=1)
        self.client.login(username="operator", password="pass12345")

        extinguishers = self.client.get(reverse("table_data", args=[self.organ.pk, "fire-extinguishers"]))
        self.assertContains(extinguishers, "<th>Дата</th>", html=True)

        fire_alarm = self.client.get(reverse("table_data", args=[self.organ.pk, "fire-alarm"]))
        self.assertContains(fire_alarm, "Подлежит оборудованию ПС")
        self.assertContains(fire_alarm, "Оборудовано ПС объектов")
        self.assertContains(fire_alarm, "Объектов с неисправной ПС")

        security_alarm = self.client.get(reverse("table_data", args=[self.organ.pk, "security-alarm"]))
        self.assertContains(security_alarm, "Подлежит оборудованию ОС")
        self.assertContains(security_alarm, "Оборудовано ОС объектов")
        self.assertContains(security_alarm, "Объектов с неисправной ОС")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "fire-alarm", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Дата")
        self.assertEqual(sheet["B1"].value, "Подлежит оборудованию ПС")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00D6EAF7")
        self.assertEqual(sheet["D2"].border.right.style, "medium")

    def test_fire_extinguisher_expiry_warning_is_cell_badge_only(self):
        today = timezone.localdate()
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date=today, required_count=10, available_count=8, expiry_date=today - timedelta(days=1), writeoff_count=1)
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date=today, required_count=10, available_count=8, expiry_date=today + timedelta(days=10), writeoff_count=1)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "fire-extinguishers"]), {"state_mode": "history"})

        self.assertContains(response, "Истек")
        self.assertContains(response, "Скоро истекает")
        self.assertContains(response, "status-rejected")
        self.assertContains(response, "status-in_work")
        self.assertNotContains(response, "row-expired")
        self.assertNotContains(response, "row-expiring")

    def test_state_snapshot_tables_show_current_records_by_default_and_history_on_request(self):
        other_organ = TerritorialOrgan.objects.create(name="Second territorial organ", order_number=2)
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date="2026-06-01", required_count=10, available_count=6, expiry_date="2026-07-10", writeoff_count=1)
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date="2026-07-01", required_count=10, available_count=8, expiry_date="2026-12-31", writeoff_count=0)
        FireExtinguisher.objects.create(territorial_organ=other_organ, state_date="2026-07-02", required_count=12, available_count=9, expiry_date="2026-08-01", writeoff_count=1)
        FireAlarm.objects.create(territorial_organ=self.organ, state_date="2026-06-01", required_objects=5, equipped_objects=3, broken_objects=2)
        FireAlarm.objects.create(territorial_organ=self.organ, state_date="2026-07-01", required_objects=5, equipped_objects=5, broken_objects=0)
        SecurityAlarm.objects.create(territorial_organ=self.organ, state_date="2026-06-01", required_objects=7, equipped_objects=4, broken_objects=2)
        SecurityAlarm.objects.create(territorial_organ=self.organ, state_date="2026-07-01", required_objects=7, equipped_objects=6, broken_objects=1)
        ServiceHousing.objects.create(territorial_organ=self.organ, state_date="2026-06-01", total_count=10, used_by_staff=4, ready_to_move=6)
        ServiceHousing.objects.create(territorial_organ=self.organ, state_date="2026-07-01", total_count=10, used_by_staff=8, ready_to_move=2)
        self.client.login(username="operator", password="pass12345")

        extinguishers = self.client.get(
            reverse("table_data", args=[self.organ.pk, "fire-extinguishers"]),
            {"organ_ids": [self.organ.pk, other_organ.pk]},
        )
        self.assertContains(extinguishers, "Последняя запись")
        self.assertContains(extinguishers, "Second territorial organ")
        self.assertContains(extinguishers, "31.12.2026")
        self.assertNotContains(extinguishers, "10.07.2026")

        history = self.client.get(
            reverse("table_data", args=[self.organ.pk, "fire-extinguishers"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "state_mode": "history"},
        )
        self.assertContains(history, "режим: История записей")
        self.assertContains(history, "10.07.2026")
        self.assertContains(history, "31.12.2026")

        for table_key, old_value, current_value in (
            ("fire-alarm", "3", "5"),
            ("security-alarm", "4", "6"),
            ("service-housing", "4", "8"),
        ):
            response = self.client.get(reverse("table_data", args=[self.organ.pk, table_key]))
            self.assertContains(response, "Последняя запись")
            self.assertContains(response, current_value)
            self.assertNotContains(response, old_value)

    def test_fire_extinguishers_can_filter_sort_and_export_by_expiry(self):
        today = timezone.localdate()
        other_organ = TerritorialOrgan.objects.create(name="Second territorial organ", order_number=2)
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date=today, required_count=10, available_count=8, expiry_date=today + timedelta(days=10), writeoff_count=1)
        FireExtinguisher.objects.create(territorial_organ=self.organ, state_date=today, required_count=5, available_count=5, expiry_date=today + timedelta(days=90), writeoff_count=0)
        FireExtinguisher.objects.create(territorial_organ=other_organ, state_date=today, required_count=12, available_count=7, expiry_date=today - timedelta(days=5), writeoff_count=2)
        FireExtinguisher.objects.create(territorial_organ=other_organ, state_date=today, required_count=6, available_count=4, expiry_date=today + timedelta(days=20), writeoff_count=0)
        self.client.login(username="operator", password="pass12345")

        grouped_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "fire-extinguishers"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "expiry_state": "soon", "expiry_order": "soonest"},
        )

        self.assertContains(grouped_response, "Second territorial organ")
        self.assertContains(grouped_response, "Скоро истекает")
        self.assertContains(grouped_response, (today + timedelta(days=20)).strftime("%d.%m.%Y"))
        self.assertNotContains(grouped_response, (today + timedelta(days=10)).strftime("%d.%m.%Y"))
        self.assertNotContains(grouped_response, (today + timedelta(days=90)).strftime("%d.%m.%Y"))
        self.assertNotContains(grouped_response, (today - timedelta(days=5)).strftime("%d.%m.%Y"))

        export_response = self.client.get(
            reverse("export_table", args=[self.organ.pk, "fire-extinguishers", "xlsx"]),
            {"organ_ids": [self.organ.pk, other_organ.pk], "expiry_state": "soon", "expiry_order": "soonest"},
        )
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet["D1"].value, "Срок годности (эксплуатации)")
        self.assertEqual(sheet["D2"].value, (today + timedelta(days=20)).strftime("%d.%m.%Y"))

    def test_fire_request_has_comment_history_filters_and_styled_export(self):
        included = FireDepartmentRequest.objects.create(territorial_organ=self.organ, request_number="F-1", request_date="2026-06-20", status="in_work", comment="Recharge")
        excluded = FireDepartmentRequest.objects.create(territorial_organ=self.organ, request_number="F-2", request_date="2026-06-20", status="done", comment="Recharge")
        self.create_status_history_entry(included)
        self.client.login(username="operator", password="pass12345")

        table_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "fire-requests"]),
            {"status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "Recharge"},
        )
        self.assertContains(table_response, "<th>Описание</th>", html=True)
        self.assertContains(table_response, "bi-clock-history")
        self.assertContains(table_response, included.request_number)
        self.assertNotContains(table_response, excluded.request_number)
        self.assertContains(table_response, "status=in_work")
        self.assertContains(table_response, "В работе")

        update_response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "fire-requests", included.pk]),
            {
                "request_number": "F-1",
                "request_date": "2026-06-20",
                "status": "done",
                "completed_at": "2026-06-29",
                "comment": "Completed",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(update_response.status_code, 200)
        history = self.status_history(included).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("fire_request_status_history", args=[self.organ.pk, included.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки F-1")
        self.assertContains(modal, "Дата исполнения заявки")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "fire-requests", "xlsx"]), {"status": "done", "q": "Completed"})
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["D1"].value, "Описание")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["D1"].border.right.style, "medium")

    def test_anti_terror_request_table_history_filters_and_styled_export(self):
        included = AntiTerrorMeasure.objects.create(territorial_organ=self.organ, request_number="A-1", request_date="2026-06-20", status="in_work", comment="Survey act")
        excluded = AntiTerrorMeasure.objects.create(territorial_organ=self.organ, request_number="A-2", request_date="2026-06-20", status="done", comment="Survey act")
        self.create_status_history_entry(included)
        self.client.login(username="operator", password="pass12345")

        table_response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "anti-terror"]),
            {"status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "Survey"},
        )
        self.assertContains(table_response, "<th>Номер</th>", html=True)
        self.assertContains(table_response, "<th>Дата</th>", html=True)
        self.assertContains(table_response, "<th>Исполнение</th>", html=True)
        self.assertContains(table_response, "<th>Описание</th>", html=True)
        self.assertNotContains(table_response, "Потребность финансирования")
        self.assertContains(table_response, "bi-clock-history")
        self.assertContains(table_response, included.request_number)
        self.assertNotContains(table_response, excluded.request_number)
        self.assertContains(table_response, "status=in_work")

        update_response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "anti-terror", included.pk]),
            {
                "request_number": "A-1",
                "request_date": "2026-06-20",
                "status": "done",
                "completed_at": "2026-06-29",
                "comment": "Completed",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(update_response.status_code, 200)
        history = self.status_history(included).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("anti_terror_status_history", args=[self.organ.pk, included.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки A-1")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "anti-terror", "xlsx"]), {"status": "done", "q": "Completed"})
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Номер")
        self.assertEqual(sheet["D1"].value, "Описание")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["D1"].border.right.style, "medium")

    def test_uoto_service_housing_has_date_without_status(self):
        ServiceHousing.objects.create(territorial_organ=self.organ, state_date="2026-06-27", total_count=10, used_by_staff=7, ready_to_move=3)
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "service-housing"]))

        self.assertContains(response, "<th>Дата</th>", html=True)
        self.assertContains(response, "<th>Общее количество</th>", html=True)
        self.assertNotContains(response, "<th>Статус</th>", html=True)
        self.assertContains(response, "table-service-housing")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "service-housing", "xlsx"]))
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Дата")
        self.assertEqual(sheet["D1"].value, "Готово к заселению")
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_business_count_validation_rejects_illogical_values(self):
        invalid_vehicle = VehicleInventory(
            territorial_organ=self.organ,
            state_date="2026-06-27",
            required_count=5,
            available_count=8,
            broken_count=1,
            writeoff_count=0,
        )
        with self.assertRaises(ValidationError):
            invalid_vehicle.full_clean()

        invalid_housing = ServiceHousing(
            territorial_organ=self.organ,
            state_date="2026-06-27",
            total_count=10,
            used_by_staff=8,
            ready_to_move=4,
        )
        with self.assertRaises(ValidationError):
            invalid_housing.full_clean()

        invalid_alarm = FireAlarm(
            territorial_organ=self.organ,
            state_date="2026-06-27",
            required_objects=4,
            equipped_objects=5,
            broken_objects=0,
        )
        with self.assertRaises(ValidationError):
            invalid_alarm.full_clean()

    def test_citsizi_quantity_must_be_positive(self):
        request_obj = CitsiziEquipment(
            territorial_organ=self.organ,
            request_number="C-0",
            request_date="2026-06-27",
            equipment_type="communication",
            quantity=0,
        )
        with self.assertRaises(ValidationError):
            request_obj.full_clean()

    def test_uoto_building_repair_nested_request_history_filters_and_export(self):
        included = BuildingRepairRequest.objects.create(territorial_organ=self.organ, request_number="B-1", request_date="2026-06-20", status="in_work", comment="Roof")
        excluded = BuildingRepairRequest.objects.create(territorial_organ=self.organ, request_number="B-2", request_date="2026-06-20", status="done", comment="Roof")
        self.create_status_history_entry(included)
        Department.objects.create(name="UOTO", slug="uoto", order_number=2)
        self.client.login(username="operator", password="pass12345")

        panel = self.client.get(reverse("department_tables", args=[self.organ.pk, "uoto"]))
        self.assertContains(panel, "Текущий ремонт зданий, помещений, сооружений")

        response = self.client.get(
            reverse("table_data", args=[self.organ.pk, "building-repair"]),
            {"status": "in_work", "date_from": "2026-06-01", "date_to": "2026-06-30", "q": "B-1"},
        )

        self.assertContains(response, "nested-table-tabs")
        self.assertContains(response, "Заявка")
        self.assertContains(response, "<th>Номер</th>", html=True)
        self.assertContains(response, "<th>Дата</th>", html=True)
        self.assertContains(response, "<th>Исполнение заявки</th>", html=True)
        self.assertContains(response, included.request_number)
        self.assertContains(response, included.comment)
        self.assertNotContains(response, excluded.request_number)
        self.assertContains(response, "bi-clock-history")

        update_response = self.client.post(
            reverse("record_update", args=[self.organ.pk, "building-repair", included.pk]),
            {
                "request_number": "B-1",
                "request_date": "2026-06-20",
                "status": "done",
                "completed_at": "2026-06-29",
                "comment": "Roof",
            },
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(update_response.status_code, 200)
        history = self.status_history(included).get(old_status="in_work", new_status="done")
        self.assertEqual(history.completed_at.isoformat(), "2026-06-29")

        modal = self.client.get(reverse("building_repair_status_history", args=[self.organ.pk, included.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(modal, "История изменений статуса заявки B-1")

        export_response = self.client.get(reverse("export_table", args=[self.organ.pk, "building-repair", "xlsx"]), {"status": "done", "q": "B-1"})
        workbook = self.response_workbook(export_response)
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Номер")
        self.assertEqual(sheet["C1"].value, "Исполнение заявки")
        self.assertEqual(sheet["D1"].value, "Описание")
        self.assertEqual(sheet["D2"].value, "Roof")
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_table_shows_only_business_fields_and_status(self):
        request_obj = TmcRequest.objects.create(
            territorial_organ=self.organ,
            created_by=self.user,
            updated_by=self.user,
            request_number="16/TMC",
            request_date="2026-06-27",
            status="in_work",
        )
        TmcRequestItem.objects.create(request=request_obj, name="Cartridge", quantity=2, unit="pcs")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("table_data", args=[self.organ.pk, "tmc-requests"]))

        self.assertContains(response, "Сведения о потребности ТМЦ")
        self.assertContains(response, "Заявка")
        self.assertContains(response, "Наименование")
        self.assertContains(response, "Количество")
        self.assertContains(response, "Номер")
        self.assertContains(response, "Дата")
        self.assertContains(response, "Исполнение заявки")
        self.assertContains(response, "Описание")
        self.assertContains(response, "16/TMC")
        self.assertContains(response, "Cartridge")
        self.assertContains(response, "2 pcs")
        self.assertContains(response, request_obj.get_status_display())
        self.assertNotContains(response, "operator")

    def test_deleted_record_disappears_from_table_for_admin(self):
        admin = get_user_model().objects.create_superuser("admin2", password="pass12345")
        UserProfile.objects.create(user=admin, role=UserProfile.Role.ADMIN)
        item = TmcRequest.objects.create(territorial_organ=self.organ, request_number="17/TMC", request_date="2026-06-27", status="in_work")
        TmcRequestItem.objects.create(request=item, name="Deleted item", quantity=1, unit="pcs")
        self.client.login(username="admin2", password="pass12345")

        response = self.client.post(reverse("record_delete", args=[self.organ.pk, "tmc-requests", item.pk]), HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        item.refresh_from_db()
        self.assertTrue(item.is_deleted)
        self.assertNotContains(response, "Deleted item")

    def test_delete_confirmation_uses_common_modal_style(self):
        item = TmcRequest.objects.create(territorial_organ=self.organ, request_number="18/TMC", request_date="2026-06-27", status="in_work")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("record_delete", args=[self.organ.pk, "tmc-requests", item.pk]), HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "delete-confirmation")
        self.assertContains(response, "Подтверждение удаления")
        self.assertContains(response, "bi-exclamation-triangle")

    def test_tmc_item_errors_use_common_modal_style(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("record_create", args=[self.organ.pk, "tmc-requests"]),
            {
                "request_number": "19/TMC",
                "request_date": "2026-06-27",
                "status": "in_work",
                "comment": "",
                "item_name": [""],
                "item_quantity": [""],
                "item_unit": ["шт."],
            },
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "modal-error-list tmc-item-errors")
        self.assertContains(response, "Добавьте хотя бы одну позицию заявки.")
        self.assertContains(response, "data-add-tmc-item")
        self.assertContains(response, "data-tmc-item-row")
        self.assertFalse(TmcRequest.objects.filter(request_number="19/TMC").exists())

    def test_department_tabs_are_separate_tables(self):
        self.assertEqual(TABLES["tmc"][0]["title"], "Заявка")
        self.assertEqual(TABLES["antiterror"][0]["title"], "Заявка (акт обследования)")
        self.assertEqual([item["key"] for item in TABLES["tmc"]], ["tmc-requests"])
        self.assertEqual([item["key"] for item in TABLES["transport"]], ["vehicle-repair", "vehicle-fuel"])
        self.assertIn("vehicle-inventory", TABLE_BY_KEY)
        self.assertEqual([item["key"] for item in TABLES["fire"]], ["fire-extinguishers", "fire-alarm", "security-alarm", "fire-requests"])
        self.assertEqual([item["key"] for item in TABLES["uoto"]], ["service-housing", "building-repair"])

    def test_photo_upload_and_soft_delete(self):
        self.client.login(username="operator", password="pass12345")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("photo.png", buffer.getvalue(), content_type="image/png")
        response = self.client.post(reverse("photo_create", args=[self.organ.pk]), {"image": image, "description": "Facade"}, HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 200)
        photo = TerritorialOrganPhoto.objects.get()
        response = self.client.post(reverse("photo_delete", args=[self.organ.pk, photo.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 200)
        photo.refresh_from_db()
        self.assertTrue(photo.is_deleted)
        self.assertTrue(AuditLog.objects.filter(action=AuditLog.Action.DELETE, model_name="TerritorialOrganPhoto").exists())

    def test_photo_form_uses_custom_single_file_picker(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_create", args=[self.organ.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(response, "data-single-file-picker")
        self.assertContains(response, "Выбрать изображение")
        self.assertContains(response, 'type="file"')

    def test_photo_edit_form_shows_preview_and_custom_folder_select(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Folder")
        photo = self.create_photo("edit-preview.png")
        photo.folder = folder
        photo.description = "Preview description"
        photo.save(update_fields=["folder", "description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_update", args=[self.organ.pk, photo.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(response, "photo-edit-preview")
        self.assertContains(response, "edit-preview.png")
        self.assertContains(response, "Preview description")
        self.assertContains(response, "photo-edit-replace")
        self.assertContains(response, "data-single-file-preview")
        self.assertContains(response, "Выбрать изображение")
        self.assertContains(response, '<i class="bi bi-image" aria-hidden="true"></i> Выбрать изображение', html=True)
        self.assertContains(response, "Parent / Folder")
        self.assertContains(response, "custom-select-field")
        self.assertContains(response, "custom-select")
        self.assertContains(response, "custom-select-native", count=0)

    def test_photo_replace_updates_upload_date(self):
        photo = self.create_photo("old-photo.png")
        old_created_at = timezone.now() - timedelta(days=3)
        TerritorialOrganPhoto.objects.filter(pk=photo.pk).update(created_at=old_created_at)
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "blue").save(buffer, format="PNG")
        image = SimpleUploadedFile("new-photo.png", buffer.getvalue(), content_type="image/png")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("photo_update", args=[self.organ.pk, photo.pk]),
            {"image": image, "description": "Updated photo"},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        photo.refresh_from_db()
        self.assertGreater(photo.created_at, old_created_at)
        self.assertEqual(photo.original_filename, "new-photo.png")

    def create_photo(self, filename="photo.png"):
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile(filename, buffer.getvalue(), content_type="image/png")
        return TerritorialOrganPhoto.objects.create(territorial_organ=self.organ, image=image, created_by=self.user, updated_by=self.user)

    def test_photo_assets_can_be_managed_only_by_author_department(self):
        User = get_user_model()
        fire_department = Department.objects.create(name="Fire", slug="fire", order_number=2)
        transport_department = Department.objects.create(name="Transport", slug="transport", order_number=3)
        fire_user = User.objects.create_user("fire-photo", password="pass12345")
        transport_user = User.objects.create_user("transport-photo", password="pass12345")
        fire_profile = UserProfile.objects.create(user=fire_user, role=UserProfile.Role.OPERATOR)
        transport_profile = UserProfile.objects.create(user=transport_user, role=UserProfile.Role.OPERATOR)
        fire_profile.allowed_departments.set([fire_department])
        transport_profile.allowed_departments.set([transport_department])
        folder = TerritorialOrganPhotoFolder.objects.create(
            territorial_organ=self.organ,
            name="Fire folder",
            created_by=fire_user,
            updated_by=fire_user,
            created_department=fire_department,
        )
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("fire-photo.png", buffer.getvalue(), content_type="image/png")
        photo = TerritorialOrganPhoto.objects.create(
            territorial_organ=self.organ,
            folder=folder,
            image=image,
            created_by=fire_user,
            updated_by=fire_user,
            created_department=fire_department,
        )

        self.client.login(username="transport-photo", password="pass12345")

        response = self.client.get(reverse("photo_update", args=[self.organ.pk, photo.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 404)
        response = self.client.post(reverse("photo_delete", args=[self.organ.pk, photo.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 404)
        response = self.client.get(reverse("photo_folder_update", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 404)
        response = self.client.post(reverse("photo_folder_delete", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 404)

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": folder.pk})
        self.assertContains(response, "fire-photo.png")
        self.assertNotContains(response, reverse("photo_update", args=[self.organ.pk, photo.pk]))
        self.assertNotContains(response, reverse("photo_folder_update", args=[self.organ.pk, folder.pk]))

        self.client.logout()
        self.client.login(username="fire-photo", password="pass12345")
        response = self.client.get(reverse("photo_update", args=[self.organ.pk, photo.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 200)
        response = self.client.get(reverse("photo_folder_update", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 200)

    def test_photo_upload_stores_author_department(self):
        fire_department = Department.objects.create(name="Fire", slug="fire", order_number=2)
        self.user.profile.allowed_departments.set([fire_department])
        self.client.login(username="operator", password="pass12345")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("department-photo.png", buffer.getvalue(), content_type="image/png")

        response = self.client.post(reverse("photo_create", args=[self.organ.pk]), {"image": image, "description": "Department photo"}, HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        photo = TerritorialOrganPhoto.objects.get(original_filename="department-photo.png")
        self.assertEqual(photo.created_by, self.user)
        self.assertEqual(photo.created_department, fire_department)

    def test_photo_upload_and_nested_folder_creation_are_denied_in_foreign_folder(self):
        User = get_user_model()
        fire_department = Department.objects.create(name="Fire", slug="fire", order_number=2)
        transport_department = Department.objects.create(name="Transport", slug="transport", order_number=3)
        fire_user = User.objects.create_user("fire-folder-owner", password="pass12345")
        fire_profile = UserProfile.objects.create(user=fire_user, role=UserProfile.Role.OPERATOR)
        self.user.profile.allowed_departments.set([transport_department])
        fire_profile.allowed_departments.set([fire_department])
        foreign_folder = TerritorialOrganPhotoFolder.objects.create(
            territorial_organ=self.organ,
            name="Foreign folder",
            created_by=fire_user,
            updated_by=fire_user,
            created_department=fire_department,
        )
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_bulk_upload", args=[self.organ.pk]), {"folder": foreign_folder.pk}, HTTP_HX_REQUEST="true")
        self.assertEqual(response.status_code, 404)

        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("foreign-folder-upload.png", buffer.getvalue(), content_type="image/png")
        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": [image], "folder": foreign_folder.pk},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(response.status_code, 404)
        self.assertFalse(TerritorialOrganPhoto.objects.filter(original_filename="foreign-folder-upload.png").exists())

        response = self.client.post(
            reverse("photo_folder_create", args=[self.organ.pk]),
            {"name": "Nested denied", "parent": foreign_folder.pk},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(response.status_code, 404)
        self.assertFalse(TerritorialOrganPhotoFolder.objects.filter(name="Nested denied").exists())

    def test_photo_cannot_be_moved_to_foreign_folder(self):
        User = get_user_model()
        fire_department = Department.objects.create(name="Fire", slug="fire", order_number=2)
        transport_department = Department.objects.create(name="Transport", slug="transport", order_number=3)
        fire_user = User.objects.create_user("fire-move-owner", password="pass12345")
        UserProfile.objects.create(user=fire_user, role=UserProfile.Role.OPERATOR).allowed_departments.set([fire_department])
        self.user.profile.allowed_departments.set([transport_department])
        foreign_folder = TerritorialOrganPhotoFolder.objects.create(
            territorial_organ=self.organ,
            name="Foreign move target",
            created_by=fire_user,
            updated_by=fire_user,
            created_department=fire_department,
        )
        photo = self.create_photo("own-photo.png")
        photo.created_department = transport_department
        photo.save(update_fields=["created_department"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("photo_update", args=[self.organ.pk, photo.pk]),
            {"folder": foreign_folder.pk, "description": "Move attempt"},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        photo.refresh_from_db()
        self.assertIsNone(photo.folder)
        self.assertNotEqual(photo.description, "Move attempt")
        self.assertContains(response, "Выберите корректный вариант")

    def test_photo_download_single_and_zip(self):
        photo = self.create_photo()
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_download", args=[self.organ.pk, photo.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment", response["Content-Disposition"])

        response = self.client.get(reverse("photos_download_all", args=[self.organ.pk]), {"download_token": "photostest"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertIn("download-ready-photostest", response.cookies)
        archive_data = b"".join(response.streaming_content)
        with zipfile.ZipFile(BytesIO(archive_data)) as archive:
            self.assertTrue(any(name.endswith(".png") for name in archive.namelist()))

    def test_photo_folder_download_includes_nested_photos(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Folder")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=folder, name="Child")
        parent_photo = self.create_photo("folder-parent-download.png")
        parent_photo.folder = folder
        parent_photo.save(update_fields=["folder"])
        child_photo = self.create_photo("folder-child-download.png")
        child_photo.folder = child
        child_photo.save(update_fields=["folder"])
        outside_photo = self.create_photo("folder-outside-marker.png")
        outside_photo.save()
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_folder_download", args=[self.organ.pk, folder.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        archive_data = b"".join(response.streaming_content)
        with zipfile.ZipFile(BytesIO(archive_data)) as archive:
            names = archive.namelist()
            self.assertTrue(any(name.endswith("folder-parent-download.png") for name in names))
            self.assertTrue(any(name.startswith("Child/") and name.endswith("folder-child-download.png") for name in names))
            self.assertFalse(any("outside-marker" in name for name in names))

    def test_photos_are_paginated_and_filterable(self):
        for index in range(25):
            photo = self.create_photo(f"photo-{index}.png")
            photo.description = f"Photo {index}"
            photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["photo_page"].object_list), 24)
        self.assertContains(response, 'data-download-preparing="Подготовка архива..."')
        self.assertContains(response, "photo-page-number")
        self.assertContains(response, 'class="pagination-jump"')
        self.assertContains(response, 'data-pagination-scroll="self"')
        self.assertContains(response, 'name="page"')
        self.assertContains(response, "page=2")
        self.assertNotContains(response, "Описание не добавлено")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"q": "photo-24", "sort": "oldest"})
        self.assertContains(response, "photo-24")

    def test_photos_search_uses_filename_not_storage_path(self):
        photo = self.create_photo("target-name.png")
        photo.description = "Filename check"
        photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"q": "target-name"})
        self.assertContains(response, "Filename check")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"q": "territorial_organs"})
        self.assertNotContains(response, "Filename check")

    def test_photo_bulk_upload_creates_folder_and_descriptions(self):
        self.client.login(username="operator", password="pass12345")
        files = []
        for name in ["bulk-1.png", "bulk-2.png"]:
            buffer = BytesIO()
            Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
            files.append(SimpleUploadedFile(name, buffer.getvalue(), content_type="image/png"))

        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": files, "descriptions": ["First photo", "Second photo"], "new_folder": "Check"},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        folder = TerritorialOrganPhotoFolder.objects.get(name="Check")
        self.assertEqual(TerritorialOrganPhoto.objects.filter(folder=folder).count(), 2)
        self.assertTrue(TerritorialOrganPhoto.objects.filter(description="First photo").exists())

    def test_photo_bulk_upload_form_has_progress_state(self):
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photo_bulk_upload", args=[self.organ.pk]), HTTP_HX_REQUEST="true")

        self.assertContains(response, "data-bulk-upload-progress")
        self.assertContains(response, "data-bulk-refresh-url")
        self.assertContains(response, "data-bulk-upload-submit")

    def test_photo_bulk_upload_batch_returns_json(self):
        self.client.login(username="operator", password="pass12345")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("batch.png", buffer.getvalue(), content_type="image/png")

        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": [image], "descriptions": ["Batch photo"]},
            HTTP_X_BULK_PHOTO_BATCH="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        self.assertEqual(response.json()["created"], 1)
        self.assertEqual(response.json()["failed"], 0)
        self.assertTrue(TerritorialOrganPhoto.objects.filter(description="Batch photo").exists())

    def test_photo_bulk_upload_accepts_more_than_300_files(self):
        self.client.login(username="operator", password="pass12345")
        files = []
        for index in range(301):
            buffer = BytesIO()
            Image.new("RGB", (1, 1), "white").save(buffer, format="PNG")
            files.append(SimpleUploadedFile(f"many-{index}.png", buffer.getvalue(), content_type="image/png"))

        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": files},
            HTTP_X_BULK_PHOTO_BATCH="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["created"], 301)
        self.assertEqual(response.json()["failed"], 0)
        self.assertEqual(TerritorialOrganPhoto.objects.count(), 301)

    def test_photo_bulk_upload_uses_current_folder(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Current")
        self.client.login(username="operator", password="pass12345")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("current-folder.png", buffer.getvalue(), content_type="image/png")

        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": [image], "descriptions": ["In current folder"], "folder": folder.pk},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(TerritorialOrganPhoto.objects.filter(folder=folder, description="In current folder").exists())

    def test_photo_folder_can_be_created_inside_folder(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(
            reverse("photo_folder_create", args=[self.organ.pk]),
            {"name": "Nested", "parent": parent.pk},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        nested = TerritorialOrganPhotoFolder.objects.get(name="Nested")
        self.assertEqual(nested.parent, parent)
        self.assertContains(response, "Nested")
        self.assertEqual(response.context["selected_folder"], parent)

    def test_photo_folder_can_be_renamed(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Old name")
        self.client.login(username="operator", password="pass12345")

        form_response = self.client.get(reverse("photo_folder_update", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(form_response, "Old name")
        self.assertContains(form_response, "Сохранить")

        response = self.client.post(
            reverse("photo_folder_update", args=[self.organ.pk, folder.pk]),
            {"name": "New name"},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        folder.refresh_from_db()
        self.assertEqual(folder.name, "New name")
        self.assertEqual(folder.parent, parent)
        self.assertContains(response, "New name")
        self.assertNotContains(response, "Old name")
        self.assertTrue(AuditLog.objects.filter(action=AuditLog.Action.UPDATE, model_name="TerritorialOrganPhotoFolder").exists())

    def test_photo_folder_can_be_moved_to_another_folder(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        target = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Target")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Child")
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Moved")
        nested = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=folder, name="Nested")
        self.client.login(username="operator", password="pass12345")

        form_response = self.client.get(reverse("photo_folder_update", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")
        self.assertContains(form_response, "Расположение")
        self.assertContains(form_response, "photo-folder-form")
        self.assertContains(form_response, "Parent / Child")
        self.assertContains(form_response, "Target")
        self.assertNotContains(form_response, "Moved / Nested")

        response = self.client.post(
            reverse("photo_folder_update", args=[self.organ.pk, folder.pk]),
            {"name": "Moved", "parent": target.pk},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        folder.refresh_from_db()
        nested.refresh_from_db()
        self.assertEqual(folder.parent, target)
        self.assertEqual(nested.parent, folder)
        self.assertEqual(response.context["selected_folder"], target)

    def test_photo_folder_delete_soft_deletes_content(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Delete me")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=folder, name="Child")
        photo = self.create_photo("folder-photo.png")
        photo.folder = folder
        photo.description = "Folder photo"
        photo.save(update_fields=["folder", "description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.post(reverse("photo_folder_delete", args=[self.organ.pk, folder.pk]), HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        folder.refresh_from_db()
        photo.refresh_from_db()
        child.refresh_from_db()
        self.assertTrue(folder.is_deleted)
        self.assertTrue(child.is_deleted)
        self.assertTrue(photo.is_deleted)
        self.assertEqual(photo.folder, folder)
        self.assertEqual(child.parent, folder)
        self.assertNotContains(response, "Folder photo")
        self.assertNotContains(response, "Delete me")

    def test_photos_show_nested_folders_in_current_folder(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Child")
        sibling = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Sibling")
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": parent.pk})

        self.assertContains(response, child.name)
        self.assertNotContains(response, sibling.name)
        self.assertEqual(list(response.context["folder_path"]), [parent])

    def test_photo_bulk_upload_can_target_nested_folder(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Child")
        self.client.login(username="operator", password="pass12345")
        buffer = BytesIO()
        Image.new("RGB", (2, 2), "white").save(buffer, format="PNG")
        image = SimpleUploadedFile("nested-folder.png", buffer.getvalue(), content_type="image/png")

        response = self.client.post(
            reverse("photo_bulk_upload", args=[self.organ.pk]),
            {"images": [image], "descriptions": ["In nested folder"], "folder": child.pk},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(TerritorialOrganPhoto.objects.filter(folder=child, description="In nested folder").exists())

    def test_photo_card_shows_clickable_full_folder_path(self):
        parent = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Parent")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=parent, name="Child")
        photo = self.create_photo("nested-path.png")
        photo.folder = child
        photo.description = "Nested path photo"
        photo.save(update_fields=["folder", "description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": child.pk})

        self.assertContains(response, "Parent")
        self.assertContains(response, "Child")
        self.assertContains(response, f"?folder={child.pk}")
        self.assertContains(response, "Nested path photo")
        self.assertContains(response, "bi-chevron-right")
        self.assertContains(response, "В папке Parent: 0 фотографий, 1 папок")
        self.assertContains(response, "В папке Child: 1 фотографий, 0 папок")

    def test_photos_root_shows_only_root_photos(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Folder")
        child = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, parent=folder, name="Child")
        inside = self.create_photo("inside-folder.png")
        inside.folder = child
        inside.description = "Inside folder"
        inside.save(update_fields=["folder", "description"])
        root = self.create_photo("root-photo.png")
        root.description = "Root photo"
        root.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]))

        self.assertContains(response, "Root photo")
        self.assertContains(response, "Folder")
        self.assertContains(response, '<article class="folder-card"', html=False)
        self.assertContains(response, 'class="folder-card-open"')
        self.assertContains(response, f"folder={folder.pk}")
        self.assertContains(response, "folder-card-counts")
        self.assertNotContains(response, '<article class="folder-card" hx-get=', html=False)
        self.assertContains(response, "Редактировать папку")
        self.assertNotContains(response, "Inside folder")
        self.assertContains(response, "всего фотографий")
        self.assertContains(response, "всего папок")
        self.assertContains(response, "В корне: 1 фотографий, 1 папок")
        self.assertContains(response, 'id="photo-search-input"')
        self.assertContains(response, "input delay:500ms from:#photo-search-input")
        self.assertContains(response, '<strong>2</strong>', count=2, html=True)

    def test_photos_filter_by_folder(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Facades")
        first = self.create_photo("facade.png")
        first.folder = folder
        first.description = "Inside folder"
        first.save(update_fields=["folder", "description"])
        second = self.create_photo("other.png")
        second.description = "Without folder"
        second.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": folder.pk})

        self.assertContains(response, "Inside folder")
        self.assertNotContains(response, "Without folder")
        self.assertNotContains(response, "active-filter-bar")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": folder.pk, "q": "inside", "sort": "oldest"})

        reset_url = f'{reverse("photos", args=[self.organ.pk])}?folder={folder.pk}'
        self.assertContains(response, "active-filter-bar")
        self.assertContains(response, "inside")
        self.assertContains(response, "oldest")
        self.assertContains(response, f'hx-get="{reset_url}"')

    def test_photos_can_prioritize_photos_before_folders(self):
        TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Folder")
        photo = self.create_photo("photo-first.png")
        photo.description = "Photo first"
        photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"order": "photos"})

        self.assertEqual(response.context["photo_item_order"], "photos")
        self.assertContains(response, "photo-order-photos")
        self.assertContains(response, "order=photos")
        self.assertContains(response, "порядок: сначала фотографии")
        self.assertContains(response, "Photo first")
        self.assertContains(response, "Folder")

    def test_photo_sort_applies_to_folders(self):
        older = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Older folder")
        newer = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Newer folder")
        TerritorialOrganPhotoFolder.objects.filter(pk=older.pk).update(created_at=timezone.now() - timedelta(days=3))
        TerritorialOrganPhotoFolder.objects.filter(pk=newer.pk).update(created_at=timezone.now())
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]))

        self.assertEqual(list(response.context["folders"])[0].name, "Newer folder")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"sort": "oldest"})

        self.assertEqual(list(response.context["folders"])[0].name, "Older folder")

    def test_photos_search_includes_folder_names(self):
        TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Assembly hall")
        TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Garage")
        photo = self.create_photo("hall-photo.png")
        photo.description = "Room photo"
        photo.save(update_fields=["description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"q": "assembly"})

        self.assertContains(response, "Assembly hall")
        self.assertNotContains(response, "Garage")
        self.assertNotContains(response, "Room photo")

    def test_photos_search_is_case_insensitive_for_cyrillic(self):
        folder = TerritorialOrganPhotoFolder.objects.create(territorial_organ=self.organ, name="Фасад")
        photo = self.create_photo("facade-demo.png")
        photo.folder = folder
        photo.description = "Фасад административного здания"
        photo.save(update_fields=["folder", "description"])
        self.client.login(username="operator", password="pass12345")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"q": "фасад"})

        self.assertContains(response, "Фасад")
        self.assertNotContains(response, "Фасад административного здания")

        response = self.client.get(reverse("photos", args=[self.organ.pk]), {"folder": folder.pk, "q": "фасад"})

        self.assertContains(response, "Фасад административного здания")


class SeedCommandTests(TestCase):
    def test_seed_is_idempotent(self):
        from django.core.management import call_command

        call_command("seed_initial_data")
        first_count = TerritorialOrgan.objects.count()
        call_command("seed_initial_data")
        self.assertEqual(TerritorialOrgan.objects.count(), first_count)
